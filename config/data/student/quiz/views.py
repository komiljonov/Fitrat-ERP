import logging
from datetime import datetime, timedelta

import openpyxl
import pandas as pd
from django.db import transaction
from django.dispatch.dispatcher import logger
from django.http import HttpResponse
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, get_object_or_404, ListAPIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .check_serializers import QuizCheckSerializer
from .models import Fill_gaps, Vocabulary, Pairs, MatchPairs, Exam, QuizGaps, Answer, ExamRegistration, ObjectiveTest, \
    Cloze_Test, ImageObjectiveTest, True_False, ExamCertificate, ExamSubject
from .models import Quiz, Question
from .serializers import QuizSerializer, QuestionSerializer, FillGapsSerializer, \
    VocabularySerializer, PairsSerializer, MatchPairsSerializer, ExamSerializer, \
    QuizGapsSerializer, AnswerSerializer, ExamRegistrationSerializer, ObjectiveTestSerializer, Cloze_TestSerializer, \
    ImageObjectiveTestSerializer, True_FalseSerializer, ExamCertificateSerializer, QuizCheckingSerializer, \
    ExamSubjectSerializer
from ..groups.models import Group
from ..homeworks.models import Homework
from ..mastering.models import Mastering
from ..shop.models import Points
from ..student.models import Student
from ..subject.models import Theme
from ...account.models import CustomUser
from ...exam_results.models import QuizResult
from ...exam_results.serializers import QuizResultSerializer


class QuizCheckAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Check Quiz Answers",
        operation_description="Submit answers to a quiz and get results.",
        request_body=QuizCheckSerializer,
        responses={200: openapi.Response(description="Quiz result summary with section breakdown.")}
    )
    def post(self, request):
        serializer = QuizCheckSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        quiz = get_object_or_404(Quiz, id=data.get("quiz_id"))
        student = Student.objects.filter(user=request.user).first()
        theme = get_object_or_404(Theme, id=data.get("theme"))




        results = {
            "details": {},
            "summary": {
                "total_questions": 0,
                "correct_count": 0,
                "wrong_count": 0,
                "ball": 0.0,
                # "section_breakdown": {}
            }
        }

        quiz_questions = QuizCheckingSerializer(quiz).data["questions"]

        for question in quiz_questions:
            qtype = question["type"]
            qid = question["id"]
            question_ids = data.get("questions", [])
            for qid in question_ids:

                question = next((q for q in quiz_questions if q["id"] == qid), None)
                if question:
                    question_data = self._prepare_question_data(question)
                    print(question_data)

            user_answer = self._find_user_answer(data, qtype, qid)

            if qtype not in results["details"]:
                results["details"][qtype] = []
            # if qtype not in results["summary"]["section_breakdown"]:
            #     results["summary"]["section_breakdown"][qtype] = {"correct": 0, "wrong": 0}

            if not user_answer:
                results["summary"]["wrong_count"] += 1
                # results["summary"]["section_breakdown"][qtype]["wrong"] += 1
                continue

            is_correct, result_data = self._check_answer(question, user_answer)

            if is_correct:
                results["summary"]["correct_count"] += 1
                # results["summary"]["section_breakdown"][qtype]["correct"] += 1
            else:
                results["summary"]["wrong_count"] += 1
                # results["summary"]["section_breakdown"][qtype]["wrong"] += 1
            results["details"][qtype].append(result_data)

        existing_results = QuizResult.objects.filter(
            quiz=quiz,
            student=student,
            questions__id__in=[q["id"] for q in quiz_questions]
        )

        results["existing_results"] = QuizResultSerializer(existing_results, many=True).data

        total = quiz.count if quiz else len(quiz_questions)
        results["summary"]["total_questions"] = total
        results["summary"]["wrong_count"] = total - results["summary"]["correct_count"]
        results["summary"]["ball"] = round((results["summary"]["correct_count"] / total * 100), 2) if total > 0 else 0.0

        self._create_mastering_record(theme, student, quiz, results["summary"]["ball"])

        return Response(results)

    def _prepare_question_data(self, question):
        qtype = question["type"]
        handler = getattr(self, f"_prepare_{qtype}", None)
        return handler(question) if handler else {"id": question["id"], "type": qtype}

    def _prepare_standard(self, question):
        return {
            "id": question["id"],
            "type": "standard",
            "text": question.get("text", {}).get("name") or question.get("question", {}).get("name"),
            "answers": [{"id": a["id"], "text": a.get("text")} for a in question.get("answers", [])],
            "correct_answer": next((a["id"] for a in question.get("answers", []) if a.get("is_correct")), None)
        }

    def _prepare_true_false(self, question):
        return {
            "id": question["id"],
            "type": "true_false",
            "question_text": question.get("question", {}).get("name"),
            "correct_answer": question.get("answer", "")
        }

    def _prepare_match_pairs(self, question):
        left_items = [p for p in question.get("pairs", []) if p.get("choice") == "Left"]
        right_items = [p for p in question.get("pairs", []) if p.get("choice") == "Right"]

        pairs = []
        for left in left_items:
            right = next((r for r in right_items if r["key"] == left["key"]), None)
            if right:
                pairs.append({
                    "left_id": left["id"],
                    "left_text": left.get("pair"),
                    "right_id": right["id"],
                    "right_text": right.get("pair")
                })

        return {
            "id": question["id"],
            "type": "match_pairs",
            "pairs": pairs
        }

    def _prepare_cloze_test(self, question):
        return {
            "id": question["id"],
            "type": "cloze_test",
            "correct_sequence": [q["name"] for q in question.get("questions", [])]
        }

    def _find_user_answer(self, data, qtype, qid):
        id_fields = {
            'standard': 'question_id',
            'match_pairs': 'match_id',
            'objective_test': 'objective_id',
            'cloze_test': 'cloze_id',
            'image_objective': 'image_objective_id',
            'true_false': 'true_false_id'
        }

        id_field = id_fields.get(qtype, 'question_id')
        return next((a for a in data.get(qtype, []) if str(a.get(id_field)) == str(qid)), None)

    def _check_answer(self, question, user_answer):
        qtype = question["type"]

        # Handle any type aliases
        if qtype in ["image_objective", "image_objective_test"]:
            qtype = "image_objective"  # Standardize to one type

        checker = getattr(self, f"check_{qtype}", None)

        if not checker:
            return False, {"error": f"Unsupported question type: {qtype}", "id": question["id"]}

        try:
            return checker(question, user_answer)
        except Exception as e:
            logger.error(f"Error in {qtype} checker: {str(e)}")
            return False, {"error": str(e), "id": question["id"]}

    def check_objective_test(self, question, user_answer):
        try:
            correct_answer = question.get("answer", "")
            if not correct_answer and "answers" in question:
                correct_answer = next(
                    (a["text"] for a in question["answers"] if a.get("is_correct")),
                    ""
                )

            user_answer_text = user_answer.get("answer_ids", "")
            is_correct = str(user_answer_text).strip().lower() == str(correct_answer).strip().lower()

            return is_correct, {
                "id": question["id"],
                # "file": question.get("file", ""),
                "question_text": question.get("question", {}).get("name"),
                "correct": is_correct,
                "user_answer": user_answer_text,
                "correct_answer": correct_answer
            }
        except Exception as e:
            logger.error(f"Error processing objective test: {str(e)}")
            return False, {
                "id": question["id"],
                "correct": False,
                "error": str(e),
                "user_answer": user_answer.get("answer_ids", ""),
                "correct_answer": "Error processing correct answer"
            }

    def check_cloze_test(self, question, user_answer):
        try:
            correct_sequence = [q["name"] for q in question.get("questions", [])][::-1]
            user_sequence = user_answer.get("word_sequence", [])
            is_correct = user_sequence == correct_sequence

            return is_correct, {
                "id": question["id"],
                #                 "file": question.get("file", ""),
                "question_text": question.get("question", {}).get("name"),
                "correct": is_correct,
                "user_answer": user_sequence,
                "correct_answer": correct_sequence
            }
        except Exception as e:
            logger.error(f"Error processing cloze test: {str(e)}")
            return False, {
                "id": question["id"],
                "correct": False,
                "error": str(e),
                "user_answer": user_answer.get("word_sequence", []),
                "correct_answer": "Error processing correct sequence"
            }

    def check_image_objective(self, question, user_answer):
        try:
            correct_answer_id = question.get("answer", "")
            user_answer_id = user_answer.get("answer", "")
            is_correct = str(user_answer_id) == str(correct_answer_id)

            return is_correct, {
                "id": question["id"],
                #                 "file": question.get("file", ""),
                "question_text": question.get("question", {}).get("name"),
                "correct": is_correct,
                "user_answer": user_answer_id,
                "correct_answer": correct_answer_id,
                "image_url": question.get("image_url", "")
            }
        except Exception as e:
            logger.error(f"Error processing image objective: {str(e)}")
            return False, {
                "id": question["id"],
                "correct": False,
                "error": str(e),
                "user_answer": user_answer.get("answer", ""),
                "correct_answer": "Error processing correct answer"
            }

    def check_standard(self, question, user_answer):
        correct_answer_id = next((a["id"] for a in question.get("answers", []) if a.get("is_correct")), None)
        user_answer_id = user_answer.get("answer_id")
        is_correct = str(user_answer_id) == str(correct_answer_id)

        print(question)

        return is_correct, {
            "id": question["id"],
            "file": question.get("file", {}),
            "question_text": question.get("text", {}).get("name"),
            "correct": is_correct,
            "user_answer": user_answer_id,
            "correct_answer": correct_answer_id
        }

    def check_true_false(self, question, user_answer):
        correct_answer = question.get("answer", "").lower()
        user_choice = user_answer.get("choice", "").lower()
        is_correct = (user_choice == correct_answer)

        return is_correct, {
            "id": question["id"],
            #             "file": question.get("file", ""),
            "question_text": question.get("question", {}).get("name"),
            "correct": is_correct,
            "user_answer": user_choice,
            "correct_answer": correct_answer
        }

    def check_match_pairs(self, question, user_answer):

        left_items = {p["key"]: p["id"] for p in question.get("pairs", []) if p.get("choice") == "Left"}
        right_items = {p["key"]: p["id"] for p in question.get("pairs", []) if p.get("choice") == "Right"}

        correct_mapping = {
            key: (left_items[key], right_items[key])
            for key in left_items if key in right_items
        }

        user_pairs = user_answer.get("pairs", [])
        all_correct = True
        pair_results = []

        for pair in user_pairs:
            left_id = str(pair.get("left_id"))
            right_id = str(pair.get("right_id"))

            is_correct = any(
                str(correct_left) == left_id and str(correct_right) == right_id
                for correct_left, correct_right in correct_mapping.values()
            )

            if not is_correct:
                all_correct = False

            pair_results.append({
                "left_id": left_id,
                "right_id": right_id,
                "is_correct": is_correct
            })

        return all_correct, {
            "id": question["id"],
            "correct": all_correct,
            #             "file": question.get("file", ""),
            "pair_results": pair_results,
            "correct_mapping": {
                key: {"left_id": left, "right_id": right}
                for key, (left, right) in correct_mapping.items()
            }
        }

    def _create_mastering_record(self, theme, student, quiz, ball):
        if not student:
            return

        try:
            mastering = Mastering.objects.create(
                theme=theme,
                student=student,
                test=quiz,
                ball=ball,
                choice="Test"
            )

            homework = Homework.objects.filter(theme=theme).first()
            if homework:
                Points.objects.create(
                    point=ball,
                    from_test=mastering,
                    from_homework=homework,
                    student=student,
                    comment=f"{homework.theme.title} test results"
                )
        except Exception:
            pass


class ExamSubjectListCreate(ListCreateAPIView):
    queryset = ExamSubject.objects.all()
    serializer_class = ExamSubjectSerializer

    def get_queryset(self):
        queryset = ExamSubject.objects.all()

        options = self.request.GET.get("options")
        lang_foreign = self.request.GET.get("lang_foreign")
        lang_national = self.request.GET.get("lang_national")

        if options:
            queryset = queryset.filter(options=options)
        if lang_foreign:
            queryset = queryset.filter(lang_foreign=lang_foreign)
        if lang_national:
            queryset = queryset.filter(lang_national=lang_national)
        return queryset

class ExamSubjectDetail(RetrieveUpdateDestroyAPIView):
    queryset = ExamSubject.objects.all()
    serializer_class = ExamSubjectSerializer

    allowed_bulk_fields = {
        "order", "has_certificate", "certificate", "certificate_expire_date"
    }

    def update(self, request, *args, **kwargs):
        if isinstance(request.data, list):
            response_data = []
            errors = []

            for item in request.data:

                item_id = item.get("id")
                if not item_id:
                    errors.append({"error": "Missing ID", "item": item})
                    continue

                try:
                    instance = ExamSubject.objects.get(id=item_id)
                except ExamSubject.DoesNotExist:
                    errors.append({"error": f"ExamSubject with id={item_id} not found."})
                    continue

                filtered_data = {
                    k: v for k, v in item.items() if k in self.allowed_bulk_fields
                }

                serializer = self.get_serializer(instance, data=filtered_data, partial=True)
                try:
                    serializer.is_valid(raise_exception=True)
                    serializer.save()
                    response_data.append(serializer.data)

                except ValidationError as ve:
                    errors.append({"id": item_id, "validation_error": ve.detail})
                except Exception as e:
                    logging.exception(f"Unexpected error while updating ExamSubject {item_id}: {e}")
                    errors.append({"id": item_id, "error": str(e)})

            return Response({
                "updated": response_data,
                "errors": errors
            }, status=status.HTTP_207_MULTI_STATUS)

        # Regular update
        return super().update(request, *args, **kwargs)



class ObjectiveTestView(ListCreateAPIView):
    queryset = ObjectiveTest.objects.all()
    serializer_class = ObjectiveTestSerializer


class Cloze_TestView(ListCreateAPIView):
    queryset = Cloze_Test.objects.all()
    serializer_class = Cloze_TestSerializer


class Cloze_TestUpdate(RetrieveUpdateDestroyAPIView):
    queryset = Cloze_Test.objects.all()
    serializer_class = Cloze_TestSerializer


class ImageCloze_TestView(ListCreateAPIView):
    queryset = ImageObjectiveTest.objects.all()
    serializer_class = ImageObjectiveTestSerializer


class True_False_TestView(ListCreateAPIView):
    queryset = True_False.objects.all()
    serializer_class = True_FalseSerializer


class True_False_TestRetriveView(RetrieveUpdateDestroyAPIView):
    queryset = True_False.objects.all()
    serializer_class = True_FalseSerializer


class QuizListCreateView(ListCreateAPIView):
    queryset = Quiz.objects.all()
    serializer_class = QuizSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_queryset(self):
        queryset = Quiz.objects.all()
        theme = self.request.GET.get("theme")
        homework = self.request.GET.get("homework")
        search = self.request.GET.get("search")

        if homework:
            queryset = queryset.filter(homework__id=homework)

        if theme:
            queryset = queryset.filter(theme__id=theme)

        if search:
            queryset = queryset.filter(title__icontains=search)

        return queryset


class QuizListPgView(ListAPIView):
    queryset = Quiz.objects.all()
    serializer_class = QuizSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    pagination_class = None

    def get_queryset(self):
        queryset = Quiz.objects.all()
        theme = self.request.GET.get("theme")
        homework = self.request.GET.get("homework")
        search = self.request.GET.get("search")

        if homework:
            queryset = queryset.filter(homework__id=homework)

        if theme:
            queryset = queryset.filter(theme__id=theme)

        if search:
            queryset = queryset.filter(title__icontains=search)

        return queryset

    def get_paginated_response(self, data):
        return Response(data)


class QuizRetrieveUpdateDestroyView(RetrieveUpdateDestroyAPIView):
    queryset = Quiz.objects.all()
    serializer_class = QuizSerializer


class AnswerListCreateView(ListCreateAPIView):
    queryset = Answer.objects.all()
    serializer_class = AnswerSerializer
    permission_classes = [IsAuthenticated]


class AnswerRetrieveUpdateDestroyView(RetrieveUpdateDestroyAPIView):
    queryset = Answer.objects.all()
    serializer_class = AnswerSerializer
    permission_classes = [IsAuthenticated]


class QuestionsListCreateView(ListCreateAPIView):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated]


class QuestionRetrieveUpdateDestroyView(RetrieveUpdateDestroyAPIView):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated]


class FillGapsView(ListCreateAPIView):
    queryset = Fill_gaps.objects.all()
    serializer_class = FillGapsSerializer
    permission_classes = [IsAuthenticated]


class FillGapsDetailsView(RetrieveUpdateDestroyAPIView):
    queryset = Fill_gaps.objects.all()
    serializer_class = FillGapsSerializer
    permission_classes = [IsAuthenticated]


class VocabularyListView(ListCreateAPIView):
    queryset = Vocabulary.objects.all()
    serializer_class = VocabularySerializer
    permission_classes = [IsAuthenticated]


class VocabularyDetailsView(RetrieveUpdateDestroyAPIView):
    queryset = Vocabulary.objects.all()
    serializer_class = VocabularySerializer
    permission_classes = [IsAuthenticated]


class LeftPairsListView(ListCreateAPIView):
    queryset = Pairs.objects.all()
    serializer_class = PairsSerializer
    permission_classes = [IsAuthenticated]


class LeftPairsDetailsView(RetrieveUpdateDestroyAPIView):
    queryset = Pairs.objects.all()
    serializer_class = PairsSerializer
    permission_classes = [IsAuthenticated]


class MatchPairsListView(ListCreateAPIView):
    queryset = MatchPairs.objects.all()
    serializer_class = MatchPairsSerializer
    permission_classes = [IsAuthenticated]


class MatchPairsDetailsView(RetrieveUpdateDestroyAPIView):
    queryset = MatchPairs.objects.all()
    serializer_class = MatchPairsSerializer
    permission_classes = [IsAuthenticated]


class ExamListView(ListCreateAPIView):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        quiz = self.request.GET.get("quiz")
        choice = self.request.GET.get("choice")
        type = self.request.GET.get("type")
        is_mandatory = self.request.GET.get("is_mandatory")
        students = self.request.GET.get("students")
        subject = self.request.GET.get("subject")
        date = self.request.GET.get("date")
        start_time = self.request.GET.get("start_time")
        end_time = self.request.GET.get("end_time")
        homework = self.request.GET.get("homework")
        is_language = self.request.GET.get("is_language")
        lang_foreign = self.request.GET.get("lang_foreign")
        lang_national = self.request.GET.get("lang_national")
        options = self.request.GET.get("options")

        queryset = Exam.objects.all()

        user = CustomUser.objects.filter(id=self.request.user.id).first()

        if user.role == "TEACHER":
            three_days_later = datetime.today() + timedelta(days=3)
            queryset = queryset.filter(date__gt=three_days_later)

        if user.role == "Student":
            two_days_later = datetime.today() + timedelta(days=3)
            queryset = queryset.filter(date__gt=two_days_later)

        if options:
            queryset = queryset.filter(options=options)
        if is_language:
            queryset = queryset.filter(is_language=is_language.capitalize())
        if lang_foreign:
            queryset = queryset.filter(lang_foreign=lang_foreign.capitalize())
        if lang_national:
            queryset = queryset.filter(lang_national=lang_national.capitalize())
        if quiz:
            queryset = queryset.filter(quiz__id=quiz)
        if choice:
            queryset = queryset.filter(choice=choice)
        if type:
            queryset = queryset.filter(type=type)
        if is_mandatory:
            queryset = queryset.filter(is_mandatory=is_mandatory.capitalize())
        if students:
            queryset = queryset.filter(students__id__in=students)
        if subject:
            queryset = queryset.filter(subject__id__in=subject)
        if date:
            queryset = queryset.filter(date=date)
        if start_time:
            queryset = queryset.filter(start_time__gte=start_time)
        if start_time and end_time:
            queryset = queryset.filter(start_time__gte=start_time, end_time__lte=end_time)
        if homework:
            queryset = queryset.filter(homework__id=homework)

        return queryset

    def perform_create(self, serializer):
        quiz = serializer.save()
        self.update_students_count(quiz)

    def update_students_count(self, quiz):
        if quiz.students_xml and quiz.students_xml.file:
            try:
                df = pd.read_excel(quiz.students_xml.file)
                quiz.students_count = len(df)
                quiz.save(update_fields=['students_count'])
            except Exception as e:
                # optionally: log the error
                print(f"Failed to parse Excel for quiz {quiz.id}: {e}")


class ExamDetailsView(RetrieveUpdateDestroyAPIView):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [IsAuthenticated]

    def perform_update(self, serializer):
        quiz = serializer.save()

        file_id = self.request.data.get("students_excel")  # ✅ this is UUID from frontend
        if file_id:
            from ...upload.models import File  # or your correct File model path

            try:
                file_instance = File.objects.get(id=file_id)
                self.update_students_count_from_file(quiz, file_instance.file)
                print("✅ Student count updated from Excel")
            except File.DoesNotExist:
                print(f"❌ File not found for ID: {file_id}")
            except Exception as e:
                print(f"❌ Error updating student count: {e}")

    def update_students_count_from_file(self, quiz, file_obj):
        try:
            df = pd.read_excel(file_obj)
            quiz.students_count = len(df)
            quiz.save(update_fields=['students_count'])
            print(f"✅ ROWS COUNTED: {len(df)}")
        except Exception as e:
            print(f"❌ Failed to parse Excel for quiz {quiz.id}: {e}")


class QuizGapsListView(ListCreateAPIView):
    queryset = QuizGaps.objects.all()
    serializer_class = QuizGapsSerializer
    permission_classes = [IsAuthenticated]


class QuizGapsDetailsView(RetrieveUpdateDestroyAPIView):
    queryset = QuizGaps.objects.all()
    serializer_class = QuizGapsSerializer
    permission_classes = [IsAuthenticated]


class ExcelQuizUploadAPIView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @swagger_auto_schema(
        operation_summary="Import Quiz Questions from Excel",
        operation_description="Upload an Excel file containing questions and answers to import them into the system.",
        manual_parameters=[
            openapi.Parameter(
                name='file',
                in_=openapi.IN_FORM,
                type=openapi.TYPE_FILE,
                required=True,
                description="Excel file (.xlsx) with a 'Questions' sheet"
            )
        ],
        responses={200: openapi.Response(description="Import result summary")}
    )
    def post(self, request):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(filename=file_obj)
            ws = wb.active

            created_questions = 0
            quiz_map = {}

            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                    quiz_title, question_text, answer_text, is_correct = row

                    if not all([quiz_title, question_text, answer_text]):
                        continue

                    # Get or create quiz
                    quiz = quiz_map.get(quiz_title)
                    if not quiz:
                        quiz, _ = Quiz.objects.get_or_create(title=quiz_title)
                        quiz_map[quiz_title] = quiz

                    # Get or create QuizGaps for question text
                    gap, _ = QuizGaps.objects.get_or_create(name=question_text)

                    # Get or create Question
                    question = Question.objects.filter(text=gap, quiz=quiz).first()
                    if not question:
                        question = Question.objects.create(text=gap, quiz=quiz)

                    # Get or create Answer
                    answer, created = Answer.objects.get_or_create(text=answer_text)
                    if created:
                        answer.is_correct = str(is_correct).strip().lower() in ["true", "1", "yes"]
                        answer.save()

                    # Add answer to question
                    question.answers.add(answer)

                    created_questions += 1

            return Response({
                "message": "Quiz imported successfully",
                "quizzes_created": len(quiz_map),
                "questions_processed": created_questions
            })

        except Exception as e:
            return Response({
                "error": str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class ExamRegistrationListCreateAPIView(ListCreateAPIView):
    queryset = ExamRegistration.objects.all()
    serializer_class = ExamRegistrationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = ExamRegistration.objects.all()

        student = self.request.GET.get("student")
        exam = self.request.GET.get("exam")
        status = self.request.GET.get("status")
        is_participating = self.request.GET.get("is_participating")
        option = self.request.GET.get("option")
        has_certificate = self.request.GET.get("has_certificate")
        group = self.request.GET.get("group")

        if group:
            qs = qs.filter(group__id=group)
        if has_certificate:
            qs = qs.filter(has_certificate=has_certificate.capitalize())
        if student:
            qs = qs.filter(student__user__id=student)
        if exam:
            qs = qs.filter(exam__id=exam)
        if status:
            qs = qs.filter(status=status)
        if is_participating:
            qs = qs.filter(is_participating=is_participating.capitalize())
        if option:
            qs = qs.filter(option=option)
        return qs


class ExamRegistrationNoPgAPIView(ListCreateAPIView):
    queryset = ExamRegistration.objects.all()
    serializer_class = ExamRegistrationSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = ExamRegistration.objects.all()

        student = self.request.GET.get("student")
        exam = self.request.GET.get("exam")
        status = self.request.GET.get("status")
        is_participating = self.request.GET.get("is_participating")
        option = self.request.GET.get("option")
        has_certificate = self.request.GET.get("has_certificate")
        group = self.request.GET.get("group")

        if group:
            qs = qs.filter(group__id=group)
        if has_certificate:
            qs = qs.filter(has_certificate=has_certificate.capitalize())
        if student:
            qs = qs.filter(student__id=student)
        if exam:
            qs = qs.filter(exam__id=exam)
        if status:
            qs = qs.filter(status=status)
        if is_participating:
            qs = qs.filter(is_participating=is_participating.capitalize())
        if option:
            qs = qs.filter(option=option)
        return qs

    def get_paginated_response(self, data):
        return Response(data)


class ExamRegisteredStudentAPIView(APIView):
    # permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @swagger_auto_schema(
        operation_summary="Export Registered Students to Excel",
        operation_description="Download an Excel file containing registered students for a specific exam.",
        manual_parameters=[
            openapi.Parameter(
                name='exam_id',
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=True,
                description="ID of the exam"
            )
        ],
        responses={200: openapi.Response(description="Excel file with registered students")}
    )
    def get(self, request):
        exam_id = request.GET.get("exam_id")
        if not exam_id:
            return Response({"detail": "Missing 'exam_id' parameter."}, status=400)

        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response({"detail": "Exam not found."}, status=404)

        registrations = ExamRegistration.objects.filter(exam=exam).select_related('student')

        # Sort by is_participating (True first)
        registrations = sorted(registrations, key=lambda r: not r.is_participating)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registered Students"

        headers = [
            "F.I.O", "Telefon raqami", "Ro'yxatdan o'tish", "Imtihonda qatnashadimi?",
            "Ball", "Talaba izohi", "Variant", "Sertifikati egasimi","Ta'lim tili"
        ]
        ws.append(headers)

        # Define fill styles
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red

        for reg in registrations:

            certificate = ExamCertificate.objects.filter(exam=exam,student=reg.student).first()
            has_certificate = False
            if certificate and certificate.certificate:
                has_certificate = True

            student = reg.student
            row_data = [
                f"{student.first_name} {student.last_name}",
                getattr(student, 'phone', ''),
                "Faol" if reg.status == "Active" else "Yakunlangan",
                "Ha" if reg.is_participating else "Yo'q",
                reg.mark,
                reg.student_comment,
                "\n".join([f"{o.subject.name} - {o.options} variant" for o in reg.option.all() if o.subject]),
                "Ha" if has_certificate else "Yo'q",
                "\n".join([
                "Uzbek" if o.lang_national else "Euro" if o.lang_foreign else "Tanlanmagan"
                for o in reg.option.all()
                ])
            ]

            row = ws.append(row_data)

            # Apply row coloring
            fill = green_fill if reg.is_participating else red_fill
            for col in range(1, len(row_data) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        # Prepare response
        response = HttpResponse(
            content_type="appl  ication/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"registered_students_exam_{exam.date}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExamCertificateAPIView(ListCreateAPIView):
    queryset = ExamCertificate.objects.all()
    serializer_class = ExamCertificateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = ExamCertificate.objects.all()

        student = self.request.GET.get("student")
        exam = self.request.GET.get("exam")
        expire_date = self.request.GET.get("expire_date")
        status = self.request.GET.get("status")

        if student:
            qs = qs.filter(student__id=student)
        if exam:
            qs = qs.filter(exam__id=exam)
        if expire_date:
            qs = qs.filter(expire_date__gte=expire_date)
        if status:
            qs = qs.filter(status=status)
        return qs



class   ExamOptionCreate(APIView):
    """
    Bulk create or update ExamRegistration entries using ordinary field objects.
    """

    def post(self, request, *args, **kwargs):
        data = request.data

        if not isinstance(data, list):
            return Response({'error': 'Body must be a list of registration objects'}, status=400)

        registrations_to_create = []
        registrations_to_update = []
        errors = []

        with transaction.atomic():
            for entry in data:
                student_id = entry.get("student")
                exam_id = entry.get("exam")
                group_id = entry.get("group")
                option = entry.get("option")

                if not student_id or not exam_id or option is None:
                    errors.append({'entry': entry, 'error': 'Missing required fields'})
                    continue

                try:
                    student = Student.objects.get(id=student_id)
                    exam = Exam.objects.get(id=exam_id)
                    group = Group.objects.get(id=group_id) if group_id else None

                    incoming_options = option if isinstance(option, list) else [option]

                    # Ensure all ExamSubject instances exist
                    existing_subjects = ExamSubject.objects.filter(id__in=incoming_options)
                    if existing_subjects.count() != len(incoming_options):
                        missing_ids = set(incoming_options) - set(str(s.id) for s in existing_subjects)
                        errors.append({'entry': entry, 'error': f"Invalid ExamSubject ID(s): {list(missing_ids)}"})
                        continue

                    existing_registration = ExamRegistration.objects.filter(
                        student=student,
                        exam=exam,
                        group=group
                    ).first()

                    if existing_registration:
                        existing_option_ids = list(existing_registration.option.values_list("id", flat=True))
                        merged_option_ids = list(set(existing_option_ids).union(set(incoming_options)))
                        registrations_to_update.append((existing_registration, merged_option_ids))
                    else:
                        registrations_to_create.append((student, exam, group, incoming_options))

                except (Student.DoesNotExist, Exam.DoesNotExist, Group.DoesNotExist) as e:
                    errors.append({'entry': entry, 'error': str(e)})
                except Exception as e:
                    errors.append({'entry': entry, 'error': f"Unexpected error: {str(e)}"})

            # Create new registrations and set M2M options
            for student, exam, group, option_ids in registrations_to_create:
                reg = ExamRegistration.objects.create(student=student, exam=exam, group=group)
                reg.option.set(option_ids)

            # Update existing registrations' options
            for reg, option_ids in registrations_to_update:
                reg.option.set(option_ids)


        print((registrations_to_create),(registrations_to_update),(errors))
        response_data = {
            'created': len(registrations_to_create),
            'updated': len(registrations_to_update),
            'errors': errors,
            'errors_count': len(errors)
        }

        return Response(response_data, status=207 if errors else 200)
