import io
from datetime import timedelta, datetime

from django.db.models import Q, F
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from django_filters.rest_framework import DjangoFilterBackend
from icecream import ic
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.generics import (
    ListCreateAPIView,
    ListAPIView,
    RetrieveUpdateDestroyAPIView,
    CreateAPIView,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Archived, Frozen
from .serializers import ArchivedSerializer, StuffArchivedSerializer, FrozenSerializer
from data.account.models import CustomUser
from data.finances.timetracker.sinx import HrPulseIntegration


class ArchivedListAPIView(ListCreateAPIView):
    queryset = Archived.objects.all()
    serializer_class = ArchivedSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend, SearchFilter, OrderingFilter)

    search_fields = ("reason",)
    filterset_fields = ("reason",)
    ordering_fields = ("reason",)

    def get_queryset(self):
        queryset = Archived.objects.all()

        get = self.request.GET.get

        lid = get("lid")
        student = get("student")
        student_stage = get("student_stage")
        lid_stage = get("lid_stage")
        creator = get("creator")
        comment = get("comment")
        is_archived = get("is_archived")
        lang = get("lang")
        subject = get("subject")
        call_operator = get("call_operator")
        sales_manager = get("sales_manager")
        service_manager = get("service_manager")
        balance_from = get("balance_from")
        balance_to = get("balance_to")
        start_date_str = get("start_date")
        end_date_str = get("end_date")
        filial = get("filial")
        debts = get("debts")
        no_debts = get("no_debts")

        if no_debts:
            # Student or lid exists and their balance is >= 100000
            queryset = queryset.filter(
                Q(student__balance__gte=100000)
                | Q(
                    lid__balance__gte=100000,
                )
            )

        if debts:
            # Student or lid exists and their balance is < 100000
            queryset = queryset.filter(
                Q(student__balance__lt=100000) | Q(lid__balance__lt=100000)
            )

        if student_stage:
            queryset = queryset.filter(student__student_stage_type=student_stage)

        if lid_stage:
            queryset = queryset.filter(lid__lid_stage_type=lid_stage)

        if filial:
            queryset = queryset.filter(
                Q(student__filial_id=filial) | Q(lid__filial_id=filial)
            )

        if is_archived:
            queryset = queryset.filter(is_archived=is_archived.capitalize())

        if lid:
            queryset = queryset.filter(lid_id=lid)

        if student:
            queryset = queryset.filter(student_id=student)

        if creator:
            queryset = queryset.filter(creator_id=creator)

        if comment:
            queryset = queryset.filter(comment__icontains=comment)

        if lang:
            queryset = queryset.filter(
                Q(student__education_lang=lang) | Q(lid__education_lang=lang)
            )

        if subject:
            queryset = queryset.filter(
                Q(student__subject_id=subject) | Q(lid__subject_id=subject)
            )

        if call_operator:
            queryset = queryset.filter(
                Q(student__call_operator_id=call_operator)
                | Q(lid__call_operator_id=call_operator)
            )

        if sales_manager:
            queryset = queryset.filter(
                Q(student__sales_manager_id=sales_manager)
                | Q(lid__sales_manager_id=sales_manager)
            )

        if service_manager:
            queryset = queryset.filter(
                Q(student__service_manager_id=service_manager)
                | Q(lid__service_manager_id=service_manager)
            )

        if balance_from or balance_to:
            student_q = Q()
            lead_q = Q()

            if balance_from:
                student_q &= Q(student__balance__gte=balance_from)
                lead_q &= Q(lid__balance__gte=balance_from)

            if balance_to:
                student_q &= Q(student__balance__lte=balance_to)
                lead_q &= Q(lid__balance__lte=balance_to)

            queryset = queryset.filter(student_q | lead_q)

        date_format = "%Y-%m-%d"
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, date_format).date()
            end_date = datetime.strptime(end_date_str, date_format).date()

            start_datetime = datetime.combine(start_date, datetime.min.time())
            end_datetime = datetime.combine(end_date, datetime.max.time())

            queryset = queryset.filter(created_at__range=(start_datetime, end_datetime))

        elif start_date_str:
            start_date = datetime.strptime(start_date_str, date_format).date()
            start_datetime = datetime.combine(start_date, datetime.min.time())
            end_datetime = datetime.combine(start_date, datetime.max.time())

            queryset = queryset.filter(created_at__range=(start_datetime, end_datetime))

        return queryset.order_by(
            Coalesce(F("student__balance"), F("lid__balance")), "-created_at"
        )


class ArchivedDetailAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Archived.objects.all()
    serializer_class = ArchivedSerializer
    permission_classes = (IsAuthenticated,)


class ListArchivedListNOPgAPIView(ListAPIView):
    queryset = Archived.objects.all()
    serializer_class = ArchivedSerializer
    permission_classes = (IsAuthenticated,)

    def get_paginated_response(self, data):
        return Response(data)


class StudentArchivedListAPIView(ListAPIView):
    queryset = Archived.objects.all()
    serializer_class = ArchivedSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        queryset = Archived.objects.all()

        id = self.request.query_params.get("id", None)
        print(id)
        if id:
            queryset = queryset.filter(Q(student_id=id) | Q(lid_id=id))
        return queryset


class StuffArchive(CreateAPIView):
    serializer_class = StuffArchivedSerializer
    permission_classes = (IsAuthenticated,)

    def create(self, request, *args, **kwargs):
        user_id = request.data.get("stuff")
        user = CustomUser.objects.filter(id=user_id).first()

        ic(user)  # Debug

        if not user:
            return Response(
                {"error": "Xodim topilmadi!"}, status=status.HTTP_404_NOT_FOUND
            )

        if user.is_archived:
            return Response(
                {"error": "Xodim allaqachon arxivlangan!"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Archive in TimeTracker
        tt = HrPulseIntegration()
        if user.second_user:
            tt_response = tt.archive_employee(user.second_user)
            ic(tt_response)  # Debug

            if not tt_response or "error" in tt_response:
                return Response(
                    {"error": "TimeTracker bilan bog'lanib bo'lmadi."},
                    status=status.HTTP_502_BAD_GATEWAY,
                )

        # Local archive
        user.is_archived = True
        user.save()

        return Response({"message": "Xodim arxivlandi!"}, status=status.HTTP_200_OK)


class FrozenListCreateList(ListCreateAPIView):
    queryset = Frozen.objects.all()
    serializer_class = FrozenSerializer

    def get_queryset(self):
        queryset = Frozen.objects.all()

        creator = self.request.GET.get("creator")
        lid = self.request.GET.get("lid")
        student = self.request.GET.get("student")

        if creator:
            queryset = queryset.filter(creator_id=creator)

        if lid:
            queryset = queryset.filter(lid_id=lid)

        if student:
            queryset = queryset.filter(student_id=student)

        return queryset


class FrozenDetailAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Frozen.objects.all()
    serializer_class = FrozenSerializer
    permission_classes = (IsAuthenticated,)


class ExportLidsExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        queryset = Archived.objects.select_related(
            "lid", "student", "creator", "comment", "filial"
        )

        # Filters
        filters = {}

        filial = request.GET.get("filial")
        is_archived = request.GET.get("is_archived", "true")
        call_operator = request.GET.get("call_operator")
        service_manager = request.GET.get("service_manager")
        sales_manager = request.GET.get("sales_manager")
        is_student = request.GET.get("is_student")
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        subject = request.GET.get("subject")
        from_balance = request.GET.get("from_balance")
        to_balance = request.GET.get("to_balance")
        debt = request.GET.get("debts")
        no_debt = request.GET.get("no_debts")
        education_lang = request.GET.get("education_lang")
        student_stage_type = request.GET.get("student_stage_type")
        lid_stage_type = request.GET.get("lid_stage_type")

        if lid_stage_type:
            queryset = queryset.filter(lid__lid_stage_type=lid_stage_type)

        if student_stage_type:
            queryset = queryset.filter(student__student_stage_type=student_stage_type)

        if education_lang:
            queryset = queryset.filter(
                Q(lid__education_lang=education_lang)
                | Q(student__education_lang=education_lang)
            )
        if call_operator:
            queryset = queryset.filter(
                Q(lid__call_operator_id=call_operator)
                | Q(student__call_operator_id=call_operator)
            )
        if sales_manager:
            queryset = queryset.filter(
                Q(lid__sales_manager_id=sales_manager)
                | Q(student__sales_manager_id=sales_manager)
            )
        if service_manager:
            queryset = queryset.filter(
                Q(lid__service_manager_id=service_manager)
                | Q(student__service_manager_id=service_manager)
            )

        if filial:
            queryset = queryset.filter(
                Q(lid__filial_id=filial) | Q(student__filial_id=filial)
            )

        if subject:
            queryset = queryset.filter(
                Q(lid__subject_id=subject) | Q(student__subject_id=subject)
            )

        if from_balance and to_balance:
            queryset = queryset.filter(
                Q(lid__balance__gte=from_balance, lid__balance__lte=to_balance)
                | Q(
                    student__balance__gte=from_balance, student__balance__lte=to_balance
                )
            )
        elif from_balance:
            queryset = queryset.filter(
                Q(lid__balance__gte=from_balance)
                | Q(student__balance__gte=from_balance)
            )

        if debt:
            queryset = queryset.filter(
                Q(lid__balance__lte=100000) | Q(student__balance__lte=100000)
            )

        if no_debt:
            queryset = queryset.filter(
                Q(lid__balance__gt=100000) | Q(student__balance__gt=100000)
            )

        date_format = "%Y-%m-%d"
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, date_format).date()
            end_date = datetime.strptime(end_date_str, date_format).date()

            start_datetime = datetime.combine(start_date, datetime.min.time())
            end_datetime = datetime.combine(end_date, datetime.max.time())

            queryset = queryset.filter(created_at__range=(start_datetime, end_datetime))

        elif start_date_str:
            start_date = datetime.strptime(start_date_str, date_format).date()
            start_datetime = datetime.combine(start_date, datetime.min.time())
            end_datetime = datetime.combine(start_date, datetime.max.time())

            queryset = queryset.filter(created_at__range=(start_datetime, end_datetime))

        # Get filtered archived objects
        archived_objects = queryset.filter(**filters)

        if is_archived:
            archived_objects = archived_objects.filter(
                is_archived=is_archived.capitalize()
            )

        if is_student:
            is_student_bool = is_student.capitalize()

            # Only filter on lid.is_student if lid exists and student is null
            archived_objects = archived_objects.filter(
                Q(
                    lid__isnull=False,
                    student__isnull=True,
                    lid__is_student=is_student_bool,
                )
            )

        # # Access restriction
        # if user.role == "CALL_OPERATOR" or getattr(user, "is_call_center", False):
        #     archived_objects = archived_objects.filter(
        #         Q(lid__filial__in=user.filial.all()) | Q(student__filial__in=user.filial.all()) |
        #         Q(lid__filial__isnull=True) | Q(student__filial__isnull=True),
        #         Q(lid__call_operator=user) | Q(student__call_operator=user) |
        #         Q(lid__call_operator__isnull=True) | Q(student__call_operator__isnull=True)
        #     )
        # else:
        #     archived_objects = archived_objects.filter(
        #         Q(lid__filial__in=user.filial.all()) | Q(student__filial__in=user.filial.all()) |
        #         Q(lid__filial__isnull=True) | Q(student__filial__isnull=True)
        #     )

        # Generate Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Arxivlar"

        headers = [
            "Ism",
            "Familiya",
            "Sharif",
            "Telefon",
            "Qo‘shimcha raqam",
            "Tug‘ilgan sana",
            "Ta’lim tili",
            "O‘quv darajasi",
            "Fan",
            "Ball",
            "Filial",
            "Marketing kanali",
            "Lid/Student bosqichi turi",
            "Lid/Student bosqichi",
            "Buyurtma bosqichi",
            "Arxivlanganmi?",
            "Muzlatilganmi?",
            "Call operator",
            "Servis menejeri",
            "Sotuv menejeri",
            "Studentmi?",
            "Balans",
        ]
        ws.append(headers)

        for archived in archived_objects:
            obj = archived.lid or archived.student

            print(obj)
            if not obj:
                continue

            row = [
                obj.first_name,
                obj.last_name or "",
                obj.middle_name or "",
                getattr(obj, "phone_number", "") or getattr(obj, "phone", ""),
                getattr(obj, "extra_number", "") or "",
                obj.date_of_birth.strftime("%Y-%m-%d") if obj.date_of_birth else "",
                getattr(obj, "education_lang", "") or "",
                getattr(obj, "edu_level", "") or "",
                obj.subject.name if getattr(obj, "subject", None) else "",
                obj.ball if hasattr(obj, "ball") else "",
                obj.filial.name if getattr(obj, "filial", None) else "",
                (
                    obj.marketing_channel.name
                    if getattr(obj, "marketing_channel", None)
                    else ""
                ),
                getattr(obj, "lid_stage_type", "")
                or getattr(obj, "student_stage_type", "")
                or "",
                getattr(obj, "lid_stages", "")
                or getattr(obj, "new_student_stages", "")
                or "",
                getattr(obj, "ordered_stages", "") or "",
                "Ha" if archived.is_archived else "Yo‘q",
                "Ha" if getattr(obj, "is_frozen", False) else "Yo‘q",
                (
                    obj.call_operator.get_full_name()
                    if getattr(obj, "call_operator", None)
                    else ""
                ),
                (
                    obj.service_manager.get_full_name()
                    if getattr(obj, "service_manager", None)
                    else ""
                ),
                (
                    obj.sales_manager.get_full_name()
                    if getattr(obj, "sales_manager", None)
                    else ""
                ),
                "Ha" if getattr(obj, "is_student", False) else "Yo‘q",
                float(getattr(obj, "balance", 0)) if hasattr(obj, "balance") else 0,
            ]
            ws.append(row)

        # Auto-size columns
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max_length + 4
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"archived_lids_{now().strftime('%Y%m%d_%H%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class LidStudentArchivedStatistics(APIView):
    def get(self, request, *args, **kwargs):
        queryset = Archived.objects.filter(is_archived=True)

        # --- Filters ---
        is_lid = request.GET.get("is_lid", "") == "true"
        education_lang = request.GET.get("education_lang")
        subject = request.GET.get("subject")
        call_operator = request.GET.get("call_operator")
        service_manager = request.GET.get("service_manager")
        sales_manager = request.GET.get("sales_manager")
        balance_from = request.GET.get("balance_from")
        balance_to = request.GET.get("balance_to")
        start_date = parse_date(request.GET.get("start_date", ""))
        end_date = parse_date(request.GET.get("end_date", ""))

        if is_lid:
            queryset = queryset.filter(lid__isnull=False)
        elif is_lid == False:
            queryset = queryset.filter(student__isnull=False)

        if education_lang:
            field = "lid__education_lang" if is_lid else "student__education_lang"
            queryset = queryset.filter(**{field: education_lang})

        if subject:
            field = "lid__subject_id" if is_lid else "student__subject_id"
            queryset = queryset.filter(**{field: subject})

        if call_operator:
            field = "lid__call_operator_id" if is_lid else "student__call_operator_id"
            queryset = queryset.filter(**{field: call_operator})

        if service_manager:
            field = (
                "lid__service_manager_id" if is_lid else "student__service_manager_id"
            )
            queryset = queryset.filter(**{field: service_manager})

        if sales_manager:
            field = "lid__sales_manager_id" if is_lid else "student__sales_manager_id"
            queryset = queryset.filter(**{field: sales_manager})

        if balance_from and balance_to:
            field_from = (
                "lid__balance_from__lte" if is_lid else "student__balance_from__lte"
            )
            field_to = "lid__balance_to__gte" if is_lid else "student__balance_to__gte"
            queryset = queryset.filter(
                **{field_from: balance_from, field_to: balance_to}
            )
        elif balance_from:
            field = "lid__balance_from__lte" if is_lid else "student__balance_from__lte"
            queryset = queryset.filter(**{field: balance_from})

        if start_date and end_date:
            queryset = queryset.filter(
                created_at__gte=start_date,
                created_at__lte=end_date + timedelta(days=1) - timedelta(seconds=1),
            )
        elif start_date:
            queryset = queryset.filter(
                created_at__gte=start_date,
                created_at__lte=start_date + timedelta(days=1) - timedelta(seconds=1),
            )

        # --- Stats ---
        all_archived = queryset.count()
        archived_lids = queryset.filter(
            lid__isnull=False, lid__lid_stage_type="NEW_LID"
        ).count()
        archived_orders = queryset.filter(
            lid__isnull=False, lid__lid_stage_type="ORDERED_LID"
        ).count()  # or another field for "order"
        archived_new_students = queryset.filter(
            student__isnull=False, student__student_stage_type="NEW_STUDENT"
        ).count()
        archived_active_students = queryset.filter(
            student__isnull=False, student__student_stage_type="ACTIVE_STUDENT"
        ).count()
        debt_lid = queryset.filter(lid__isnull=False, lid__balance__gte=100000).count()
        deb_student = queryset.filter(
            student__isnull=False, student__balance__gte=100000
        ).count()

        no_debt_lid = queryset.filter(
            lid__isnull=False, lid__balance__lte=100000
        ).count()
        no_debt_student = queryset.filter(
            student__isnull=False, student__balance__lte=100000
        ).count()

        return Response(
            {
                "debt": debt_lid + deb_student,
                "no_debt": no_debt_lid + no_debt_student,
                "all_archived": all_archived,
                "archived_lids": archived_lids,
                "archived_orders": archived_orders,
                "archived_new_students": archived_new_students,
                "archived_active_students": archived_active_students,
            }
        )
