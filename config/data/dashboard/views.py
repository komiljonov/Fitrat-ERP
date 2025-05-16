from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl
from django.db.models import Case, When
from django.db.models import Count
from django.db.models import Sum, F, DecimalField, Value
from django.db.models.functions import ExtractWeekDay, Concat
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from icecream import ic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from data.department.marketing_channel.models import MarketingChannel
from data.finances.finance.models import Finance, Casher, Kind, SaleStudent, VoucherStudent
from data.lid.new_lid.models import Lid
from data.student.groups.models import Room, Group, Day
from data.student.studentgroup.models import StudentGroup
from ..account.models import CustomUser
from ..finances.compensation.models import Monitoring, Asos, Comments, StudentCatchingMonitoring, \
    Monitoring5, StudentCountMonitoring
from ..lid.new_lid.serializers import LidSerializer
from ..results.models import Results
from ..student.attendance.models import Attendance
from ..student.lesson.models import FirstLLesson
from ..student.student.models import Student
from ..upload.serializers import FileUploadSerializer


class DashboardView(APIView):
    def get(self, request, *args, **kwargs):
        # Get Query Parameters
        start_date = request.query_params.get('start_date',None)
        end_date = request.query_params.get('end_date',None)
        channel_id = request.query_params.get('marketing_channel')
        service_manager = request.query_params.get('service_manager')
        call_operator = request.query_params.get('call_operator')
        sales_manager = request.query_params.get('sales_manager')
        filial = request.query_params.get('filial' , None)
        subjects = request.query_params.get('subject')
        course = request.query_params.get('course')
        teacher = request.query_params.get('teacher')
        is_student = request.query_params.get('is_student')

        # Common Filters
        filters = {}
        if start_date and end_date is "":
            filters["created_at__date"] = start_date
        if start_date and end_date:
            filters["created_at__gte","created_at__lte"] = start_date,end_date
        if filial:
            filters["filial"] = filial

        # Initial QuerySets
        lid = Lid.objects.filter(**filters)
        archived_lid = lid.filter(lid_stage_type="NEW_LID",is_archived=True)
        orders = lid.filter(lid_stage_type="ORDERED_LID")
        orders_archived = orders.filter(is_archived=True)
        first_lesson = FirstLLesson.objects.filter(**filters)

        # Students with One Attendance
        # students_with_one_attendance = Attendance.objects.values("student").annotate(
        #     count=Count("id")).filter(count=1, **filters).values_list("student", flat=True)

        first_lesson_come = Student.objects.filter(student_stage_type="NEW_STUDENT", **filters)
        first_lesson_come_archived = first_lesson_come.filter(is_archived=True)

        # # First Course Payment Students
        # payment_students = Finance.objects.filter(
        #     student__isnull=False, kind__name="COURSE_PAYMENT", **filters
        # ).values_list("student", flat=True)

        first_course_payment = Student.objects.filter(student_stage_type="ACTIVE_STUDENT", **filters)
        first_course_payment_archived = first_course_payment.filter(is_archived=True)

        # Active and Ended Courses
        new_student = StudentGroup.objects.filter(student__student_stage_type="NEW_STUDENT", **filters)
        active_student = StudentGroup.objects.filter(student__student_stage_type="ACTIVE_STUDENT",group__status="ACTIVE", **filters)
        course_ended = StudentGroup.objects.filter(group__status="INACTIVE", **filters)

        # **Filtering Based on Dynamic Conditions**

        if is_student:
            is_student_value = is_student.capitalize()

            lid = lid.filter(is_student=is_student_value, is_archived=False)
            archived_lid = archived_lid.filter(is_student=is_student_value)
            orders = orders.filter(is_student=is_student_value,is_archived=False)
            orders_archived = orders_archived.filter(is_student=is_student_value)
            first_lesson = first_lesson.filter(lid__is_student=is_student_value, lid__is_archived=False)
            first_lesson_come = first_lesson_come.filter(is_archived=False)

            first_lesson_come_archived = first_lesson_come.filter(
                is_archived=True,is_student=False) if first_lesson_come.exists() else None
            first_course_payment = first_course_payment.filter(is_archived=is_student_value)
            first_course_payment_archived = first_course_payment.filter(
                is_archived=True) if first_course_payment.exists() else None

        if channel_id:
            channel = MarketingChannel.objects.get(id=channel_id)
            lid = lid.filter(marketing_channel=channel)
            archived_lid = archived_lid.filter(marketing_channel=channel)
            orders = orders.filter(marketing_channel=channel)
            orders_archived = orders_archived.filter(marketing_channel=channel)
            first_lesson = first_lesson.filter(lid__marketing_channel=channel)
            first_lesson_come = first_lesson_come.filter(marketing_channel=channel)
            first_lesson_come_archived = first_lesson_come_archived.filter(marketing_channel=channel)
            first_course_payment = first_course_payment.filter(marketing_channel=channel)
            first_course_payment_archived = first_course_payment_archived.filter(marketing_channel=channel)

        if service_manager:
            lid = lid.filter(service_manager_id=service_manager)
            archived_lid = archived_lid.filter(service_manager_id=service_manager)
            orders = orders.filter(service_manager_id=service_manager)
            orders_archived = orders_archived.filter(service_manager_id=service_manager)
            first_lesson = first_lesson.filter(lid__service_manager_id=service_manager)
            first_lesson_come = first_lesson_come.filter(service_manager_id=service_manager)
            first_lesson_come_archived = first_lesson_come_archived.filter(service_manager_id=service_manager)
            first_course_payment = first_course_payment.filter(service_manager_id=service_manager)
            first_course_payment_archived = first_course_payment_archived.filter(service_manager_id=service_manager)

        if sales_manager:
            lid = lid.filter(sales_manager_id=sales_manager)
            archived_lid = archived_lid.filter(sales_manager_id=sales_manager)
            orders = orders.filter(sales_manager_id=sales_manager)
            orders_archived = orders_archived.filter(sales_manager_id=sales_manager)
            first_lesson = first_lesson.filter(lid__sales_manager_id=sales_manager)
            first_lesson_come = first_lesson_come.filter(sales_manager_id=sales_manager)
            first_lesson_come_archived = first_lesson_come_archived.filter(sales_manager_id=sales_manager)
            first_course_payment = first_course_payment.filter(sales_manager_id=sales_manager)
            first_course_payment_archived = first_course_payment_archived.filter(sales_manager_id=sales_manager)

        if call_operator:
            lid = lid.filter(call_operator_id=call_operator)
            archived_lid = archived_lid.filter(call_operator_id=call_operator)
            orders = orders.filter(call_operator_id=call_operator)
            orders_archived = orders_archived.filter(call_operator_id=call_operator)
            first_lesson = first_lesson.filter(lid__call_operator_id=call_operator)
            first_lesson_come = first_lesson_come.filter(call_operator_id=call_operator)
            first_lesson_come_archived = first_lesson_come_archived.filter(call_operator_id=call_operator)
            first_course_payment = first_course_payment.filter(call_operator_id=call_operator)
            first_course_payment_archived = first_course_payment_archived.filter(call_operator_id=call_operator)

        if subjects:
            lid = lid.filter(subject_id=subjects)
            archived_lid = archived_lid.filter(subject_id=subjects)
            orders = orders.filter(subject_id=subjects)
            orders_archived = orders_archived.filter(subject_id=subjects)
            first_lesson = first_lesson.filter(lid__subject_id=subjects)
            first_lesson_come = first_lesson_come.filter(subject_id=subjects)
            first_lesson_come_archived = first_lesson_come_archived.filter(subject_id=subjects)
            first_course_payment = first_course_payment.filter(subject_id=subjects)
            first_course_payment_archived = first_course_payment_archived.filter(subject_id=subjects)

        if teacher:
            lid = lid.filter(lids_group__group__teacher_id=teacher)
            archived_lid = archived_lid.filter(lids_group__group__teacher_id=teacher)
            orders = orders.filter(lids_group__group__teacher_id=teacher)
            orders_archived = orders_archived.filter(lids_group__group__teacher_id=teacher)
            first_lesson = first_lesson.filter(group__teacher_id=teacher)
            first_lesson_come = first_lesson_come.filter(students_group__group__teacher_id=teacher)
            first_lesson_come_archived = first_lesson_come_archived.filter(students_group__group_id=teacher)
            first_course_payment = first_course_payment.filter(students_group__group__teacher_id=teacher)
            first_course_payment_archived = first_course_payment_archived.filter(students_group__group__teacher_id=teacher)

        if course:
            lid = lid.filter(lids_group__group__course_id=course)
            archived_lid = archived_lid.filter(lids_group__group__course_id=course)
            orders = orders.filter(lids_group__group__course_id=course)
            orders_archived = orders_archived.filter(lids_group__group__course_id=course)
            first_lesson = first_lesson.filter(group__course_id=course)
            first_lesson_come = first_lesson_come.filter(students_group__group__course_id=course)
            first_lesson_come_archived = first_lesson_come_archived.filter(students_group__group__course_id=course)
            first_course_payment = first_course_payment.filter(students_group__group__course_id=course)
            first_course_payment_archived = first_course_payment_archived.filter(students_group__group__course_id=course)

        # Final Data Output
        data = {
            "lids": lid.count(),
            "archived_lid": archived_lid.count(),
            "orders": orders.count(),
            "orders_archived": orders_archived.count(),
            "first_lesson": first_lesson.count(),
            "first_lesson_come": first_lesson_come.count(),
            "first_lesson_come_archived": first_lesson_come_archived.count() if first_lesson_come_archived else 0,
            "first_course_payment": first_course_payment.count(),
            "new_student": new_student.count(),
            "active_student": active_student.count(),
            "first_course_payment_archived": first_course_payment_archived.count() if first_course_payment_archived else 0,
            "course_ended": course_ended.count(),
        }

        return Response(data)


class DashboardSecondView(APIView):
    def get(self, request, *args, **kwargs):
        # Get Query Parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        channel_id = request.query_params.get('marketing_channel')
        service_manager = request.query_params.get('service_manager')
        call_operator = request.query_params.get('call_operator')
        sales_manager = request.query_params.get('sales_manager')
        filial = request.query_params.get('filial')
        subjects = request.query_params.get('subject')
        course = request.query_params.get('course')
        teacher = request.query_params.get('teacher')
        is_student = request.query_params.get('is_student')

        # Common Filters
        filters = {}
        if start_date:
            filters["created_at__gte"] = start_date
        if end_date:
            filters["created_at__lte"] = end_date
        if filial:
            filters["filial_id"] = filial

        # Initial QuerySets
        lid = Lid.objects.filter(lid_stage_type="NEW_LID",**filters).exclude(ordered_stages="BIRINCHI_DARS_BELGILANGAN")
        archived_lid = lid.filter(lid_stage_type="NEW_LID",is_archived=True,)
        orders = Lid.objects.filter(lid_stage_type="ORDERED_LID",**filters).exclude(ordered_stages="BIRINCHI_DARS_BELGILANGAN")
        orders_archived = orders.filter(is_archived=True)
        first_lesson = FirstLLesson.objects.filter(**filters)

        # Students with One Attendance
        students_with_one_attendance = Attendance.objects.values("student").annotate(
            count=Count("id")).filter(count=1, **filters).values_list("student", flat=True)

        first_lesson_come = Student.objects.filter(id__in=students_with_one_attendance, **filters)
        first_lesson_come_archived = first_lesson_come.filter(is_archived=True)

        # First Course Payment Students
        payment_students = Finance.objects.filter(
            student__isnull=False, kind__name="COURSE_PAYMENT", **filters
        ).values_list("student", flat=True)

        first_course_payment = Student.objects.filter(id__in=payment_students, **filters)
        first_course_payment_archived = first_course_payment.filter(is_archived=True)

        # Active and Ended Courses
        new_student = Student.objects.filter(student_stage_type="NEW_STUDENT", **filters)
        active_student = Student.objects.filter(student_stage_type="ACTIVE_STUDENT", **filters)
        course_ended = StudentGroup.objects.filter(group__status="INACTIVE", **filters)
        all_students = Student.objects.filter(is_archived=False)

        # **Filtering Based on Dynamic Conditions**

        if is_student:
            is_student_value = is_student.capitalize()

            lid = lid.filter(is_student=is_student_value, is_archived=False)
            archived_lid = archived_lid.filter(is_student=is_student_value)
            orders = orders.filter(is_student=is_student_value,is_archived=False)
            orders_archived = orders_archived.filter(is_student=is_student_value)
            first_lesson = first_lesson.filter(lid__is_student=is_student_value, lid__is_archived=False)
            first_lesson_come = first_lesson_come.filter(is_archived=False)

            first_lesson_come_archived = first_lesson_come.filter(
                is_archived=True) if first_lesson_come.exists() else None
            first_course_payment = first_course_payment.filter(is_archived=is_student_value)
            first_course_payment_archived = first_course_payment.filter(
                is_archived=True) if first_course_payment.exists() else None

        if channel_id:
            channel = MarketingChannel.objects.get(id=channel_id)
            lid = lid.filter(marketing_channel=channel)
            archived_lid = archived_lid.filter(marketing_channel=channel)
            orders = orders.filter(marketing_channel=channel)
            orders_archived = orders_archived.filter(marketing_channel=channel)
            first_lesson = first_lesson.filter(lid__marketing_channel=channel)
            first_lesson_come = first_lesson_come.filter(marketing_channel=channel)
            first_lesson_come_archived = first_lesson_come_archived.filter(marketing_channel=channel)
            first_course_payment = first_course_payment.filter(marketing_channel=channel)
            first_course_payment_archived = first_course_payment_archived.filter(marketing_channel=channel)

        if service_manager:
            lid = lid.filter(service_manager_id=service_manager)
            archived_lid = archived_lid.filter(service_manager_id=service_manager)
            orders = orders.filter(service_manager_id=service_manager)
            orders_archived = orders_archived.filter(service_manager_id=service_manager)
            first_lesson = first_lesson.filter(lid__service_manager_id=service_manager)
            first_lesson_come = first_lesson_come.filter(service_manager_id=service_manager)
            first_lesson_come_archived = first_lesson_come_archived.filter(service_manager_id=service_manager)
            first_course_payment = first_course_payment.filter(service_manager_id=service_manager)
            first_course_payment_archived = first_course_payment_archived.filter(service_manager_id=service_manager)

        if sales_manager:
            lid = lid.filter(sales_manager_id=sales_manager)
            archived_lid = archived_lid.filter(sales_manager_id=sales_manager)
            orders = orders.filter(sales_manager_id=sales_manager)
            orders_archived = orders_archived.filter(sales_manager_id=sales_manager)
            first_lesson = first_lesson.filter(lid__sales_manager_id=sales_manager)
            first_lesson_come = first_lesson_come.filter(sales_manager_id=sales_manager)
            first_lesson_come_archived = first_lesson_come_archived.filter(sales_manager_id=sales_manager)
            first_course_payment = first_course_payment.filter(sales_manager_id=sales_manager)
            first_course_payment_archived = first_course_payment_archived.filter(sales_manager_id=sales_manager)

        if call_operator:
            lid = lid.filter(call_operator_id=call_operator)
            archived_lid = archived_lid.filter(call_operator_id=call_operator)
            orders = orders.filter(call_operator_id=call_operator)
            orders_archived = orders_archived.filter(call_operator_id=call_operator)
            first_lesson = first_lesson.filter(lid__call_operator_id=call_operator)
            first_lesson_come = first_lesson_come.filter(call_operator_id=call_operator)
            first_lesson_come_archived = first_lesson_come_archived.filter(call_operator_id=call_operator)
            first_course_payment = first_course_payment.filter(call_operator_id=call_operator)
            first_course_payment_archived = first_course_payment_archived.filter(call_operator_id=call_operator)

        if subjects:
            lid = lid.filter(subject_id=subjects)
            archived_lid = archived_lid.filter(subject_id=subjects)
            orders = orders.filter(subject_id=subjects)
            orders_archived = orders_archived.filter(subject_id=subjects)
            first_lesson = first_lesson.filter(lid__subject_id=subjects)
            first_lesson_come = first_lesson_come.filter(subject_id=subjects)
            first_lesson_come_archived = first_lesson_come_archived.filter(subject_id=subjects)
            first_course_payment = first_course_payment.filter(subject_id=subjects)
            first_course_payment_archived = first_course_payment_archived.filter(subject_id=subjects)

        if teacher:
            lid = lid.filter(lids_group__group__teacher_id=teacher)
            archived_lid = archived_lid.filter(lids_group__group__teacher_id=teacher)
            orders = orders.filter(lids_group__group__teacher_id=teacher)
            orders_archived = orders_archived.filter(lids_group__group__teacher_id=teacher)
            first_lesson = first_lesson.filter(group__teacher_id=teacher)
            first_lesson_come = first_lesson_come.filter(students_group__group__teacher_id=teacher)
            first_lesson_come_archived = first_lesson_come_archived.filter(students_group__group_id=teacher)
            first_course_payment = first_course_payment.filter(students_group__group__teacher_id=teacher)
            first_course_payment_archived = first_course_payment_archived.filter(students_group__group__teacher_id=teacher)

        if course:
            lid = lid.filter(lids_group__group__course_id=course)
            archived_lid = archived_lid.filter(lids_group__group__course_id=course)
            orders = orders.filter(lids_group__group__course_id=course)
            orders_archived = orders_archived.filter(lids_group__group__course_id=course)
            first_lesson = first_lesson.filter(group__course_id=course)
            first_lesson_come = first_lesson_come.filter(students_group__group__course_id=course)
            first_lesson_come_archived = first_lesson_come_archived.filter(students_group__group__course_id=course)
            first_course_payment = first_course_payment.filter(students_group__group__course_id=course)
            first_course_payment_archived = first_course_payment_archived.filter(students_group__group__course_id=course)

        # Final Data Output
        data = {
            "lids": lid.count(),
            "archived_lid": archived_lid.count(),
            "archive_lid_res" : LidSerializer(lid, many=True).data,
            "orders": orders.count(),
            "orders_archived": orders_archived.count(),
            "first_lesson": first_lesson.count(),
            "first_lesson_come": first_lesson_come.count(),
            "first_lesson_come_archived": first_lesson_come_archived.count() if first_lesson_come_archived else 0,
            "first_course_payment": first_course_payment.count(),
            "new_student": new_student.count(),
            "active_student": active_student.count(),
            "first_course_payment_archived": first_course_payment_archived.count() if first_course_payment_archived else 0,
            "course_ended": course_ended.count(),
            "all_students": all_students.count(),
        }

        return Response(data)


class MarketingChannels(APIView):
    def get(self, request, *args, **kwargs):
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        filial = self.request.query_params.get('filial')

        filters = {}
        if start_date:
            filters['created_at__gte'] = start_date
        if end_date:
            filters['created_at__lte'] = end_date
        if filial:
            filters['filial'] = filial

        channels = MarketingChannel.objects.all()
        channel_counts = {}

        for channel in channels:
            channel_counts[channel.name] = Lid.objects.filter(
                marketing_channel=channel,
                **filters,
            ).count()

        return Response(channel_counts)


class CheckRoomFillingView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        filial = request.query_params.get('filial', None)
        room_id = request.query_params.get('room', None)
        lesson_type = request.query_params.get('lesson_type', None)
        start_time = request.query_params.get('start_time', None)
        end_time = request.query_params.get('end_time', None)
        lesson_duration = request.query_params.get('lesson_duration', "90")  # Default 90 min

        if not filial or not start_time or not end_time or not lesson_type:
            return Response({'error': 'Missing required parameters'}, status=400)

        try:
            start_time = datetime.strptime(start_time, "%H:%M").time()
            end_time = datetime.strptime(end_time, "%H:%M").time()
            lesson_duration = int(lesson_duration)
        except ValueError:
            return Response({'error': 'Invalid input format'}, status=400)

        # Get all rooms in the filial if no specific room is given
        if not room_id:
            room_ids = list(Room.objects.filter(filial_id=filial).values_list("id", flat=True))
        else:
            room_ids = [room_id]

        # Map lesson_type to days
        lesson_days_map = {
            "1": ["Dushanba"],
            "0": ["Seshanba"],
            ".": ["Dushanba", "Seshanba"]
        }
        lesson_days = lesson_days_map.get(lesson_type, [])

        # Get lesson days IDs
        lesson_days_ids = Day.objects.filter(name__in=lesson_days).values_list("id", flat=True)

        # Fetch all active lessons within the specified rooms and lesson days
        active_lessons = Group.objects.filter(
            room_number_id__in=room_ids,
            scheduled_day_type__id__in=lesson_days_ids,
            started_at__lt=end_time,
            ended_at__gt=start_time,
            status="ACTIVE"
        ).order_by("started_at").distinct()

        # **Statistics Calculation**
        total_available_time = (datetime.combine(datetime.today(), end_time) -
                                datetime.combine(datetime.today(), start_time)).seconds // 60
        ic(total_available_time)
        total_available_lesson_hours = (total_available_time // lesson_duration)
        ic(total_available_lesson_hours)

        # **Count occupied lesson hours**
        occupied_lesson_hours = sum(
            (lesson.ended_at.hour * 60 + lesson.ended_at.minute) -
            (lesson.started_at.hour * 60 + lesson.started_at.minute)
            for lesson in active_lessons
        ) // lesson_duration

        ic(occupied_lesson_hours)

        free_lesson_hours = total_available_lesson_hours - occupied_lesson_hours
        ic(free_lesson_hours)

        # **Count number of lesson hour pairs**
        lesson_hour_pairs = occupied_lesson_hours // 2
        ic(lesson_hour_pairs,occupied_lesson_hours)

        # **Count unique groups running in this period**
        total_groups = active_lessons.count()
        ic(total_groups)


        rooms = Room.objects.filter(filial_id=self.request.user.filial.first())


        total_students_capacity = sum(
            [room.room_filling * total_available_lesson_hours for room in rooms]
        )

        for room in rooms:
            print(room.room_filling, total_available_lesson_hours)

        # total_students_capacity = sum(
        #     (total_available_lesson_hours) * (i.room_number.room_filling if i.room_number else 0)
        #     for i in active_lessons
        # )

        ic(total_students_capacity)

        # total_students_capacity -= sum(occupied_lesson_hours * room.room_filling for room in rooms)

        ic(total_students_capacity)

        if lesson_type == "1":
            groups_students = StudentGroup.objects.filter(filial_id=filial,
                                                          group__scheduled_day_type__name__in=["Dushanba"])
        elif lesson_type == "0":
            groups_students = StudentGroup.objects.filter(filial_id=filial,
                                                          group__scheduled_day_type__name__in=["Seshanba"])
        else :
            groups_students = StudentGroup.objects.filter(filial_id=filial)


        if lesson_type == "1":
            new_students = StudentGroup.objects.filter(filial_id=filial,student__student_stage_type="NEW_STUDENT",
                                                          group__scheduled_day_type__name__in=["Dushanba"])
        elif lesson_type == "0":
            new_students = StudentGroup.objects.filter(filial_id=filial,student__student_stage_type="NEW_STUDENT",
                                                          group__scheduled_day_type__name__in=["Seshanba"])
        else :
            new_students = StudentGroup.objects.filter(filial_id=filial,student__student_stage_type="NEW_STUDENT",)

        # **Generate final statistics response**
        return Response({
            "groups_students": groups_students.count(),
            "new_students": new_students.count(),
            "total_available_lesson_hours": total_available_lesson_hours,
            "occupied_lesson_hours": occupied_lesson_hours,
            "free_lesson_hours": free_lesson_hours,
            "lesson_hour_pairs": lesson_hour_pairs,
            "total_groups": total_groups,
            "total_students_capacity": (total_students_capacity *( 1 if lesson_type in ["1", "0"] else 3)) - groups_students.count(),
            "all_places":total_students_capacity,
            "weeks_capacity": total_students_capacity *(1 if lesson_type in ["1", "0"] else 3) ,
            "AAAAAAAAAAAA": lesson_type in ["1", "0"]
        })


class DashboardLineGraphAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        from datetime import datetime
        from django.db.models import Sum
        from django.db.models.functions import ExtractWeekDay

        # Extract parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        casher_id = request.query_params.get('cashier')
        filial = request.query_params.get('filial')

        filters = {}

        # Filter by Casher if provided
        if casher_id:
            try:
                casher = Casher.objects.get(id=casher_id)
                filters['casher__id'] = casher.id
            except Casher.DoesNotExist:
                return Response({"error": "Casher not found"}, status=status.HTTP_400_BAD_REQUEST)

        # Filter by start_date and end_date if provided
        if start_date:
            try:
                filters['created_at__gte'] = datetime.strptime(start_date, '%Y-%m-%d')
            except ValueError:
                return Response({"error": "Invalid start_date format. Use YYYY-MM-DD."},
                                status=status.HTTP_400_BAD_REQUEST)

        if filial:
            filters['filial__id'] = filial

        if end_date:
            try:
                filters['created_at__lte'] = datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return Response({"error": "Invalid end_date format. Use YYYY-MM-DD."},
                                status=status.HTTP_400_BAD_REQUEST)

        # Define weekdays mapping
        weekday_map = {
            1: 'Sunday', 2: 'Monday', 3: 'Tuesday', 4: 'Wednesday',
            5: 'Thursday', 6: 'Friday', 7: 'Saturday'
        }

        # Initialize default structure for all weekdays
        full_week_data = {day: {"income": 0, "expense": 0} for day in weekday_map.values()}

        # Fetch income data
        income_queryset = (
            Finance.objects
            .filter(action="INCOME", **filters)
            .annotate(weekday=ExtractWeekDay('created_at'))
            .values('weekday')
            .annotate(total_amount=Sum('amount'))
        )

        # Fetch expense data
        expense_queryset = (
            Finance.objects
            .filter(action="EXPENSE", **filters)
            .annotate(weekday=ExtractWeekDay('created_at'))
            .values('weekday')
            .annotate(total_amount=Sum('amount'))
        )

        # Aggregate income data
        for item in income_queryset:
            weekday_name = weekday_map[item['weekday']]
            full_week_data[weekday_name]["income"] = item['total_amount']

        # Aggregate expense data
        for item in expense_queryset:
            weekday_name = weekday_map[item['weekday']]
            full_week_data[weekday_name]["expense"] = item['total_amount']

        # Compute total income and expense
        total_income = sum(item["income"] for item in full_week_data.values())
        total_expense = sum(item["expense"] for item in full_week_data.values())

        # Convert dictionary to list format
        result = [
            {
                "weekday": day,
                "income": total["income"],
                "expense": total["expense"]
            }
            for day, total in full_week_data.items()
        ]

        return Response({
            "data": result,
            "total_income": total_income,
            "total_expense": total_expense
        }, status=status.HTTP_200_OK)


class MonitoringView(APIView):
    def get(self, request, *args, **kwargs):
        # Get query parameters
        full_name = request.query_params.get('search', None)
        start_date = request.query_params.get('start_date', None)
        end_date = request.query_params.get('end_date', None)
        subject_id = request.query_params.get('subject', None)
        course_id = request.query_params.get('course', None)
        filial = request.query_params.get('filial', None)

        # Base queryset for teachers
        teachers = CustomUser.objects.filter(role__in=["TEACHER", "ASSISTANT"]).annotate(
            name=Concat(F('first_name'), Value(' '), F('last_name')),
            overall_point=F('monitoring')
        )

        if course_id:
            teachers = teachers.filter(teachers_groups__course__id=course_id)  # ✅ Correct related_name usage

        if subject_id:
            teachers = teachers.filter(teachers_groups__course__subject__id=subject_id)
        if full_name:
            teachers = teachers.filter(name__icontains=full_name)
        if filial:
            teachers = teachers.filter(filial__id=filial)
        if start_date and end_date:
            teachers = teachers.filter(created_at__date__gte=start_date, created_at__lte=end_date)
        elif start_date:
            teachers = teachers.filter(created_at__date__gte=start_date)
        elif end_date:
            teachers = teachers.filter(created_at__date__lte=end_date)

        teacher_data = []

        for teacher in teachers:
            subjects_query = (
                Group.objects.filter(teacher=teacher)
                .annotate(
                    subject_id=F("course__subject__id"),
                    subject_name=F("course__subject__name")
                )
                .values("subject_id", "subject_name")  # Ensure renamed fields are used in values()
            )
            if subject_id:
                subjects_query = subjects_query.filter(id=subject_id)

            subjects = list(subjects_query)

            if start_date and end_date:
                results = Results.objects.filter(teacher=teacher, created_at__gte=start_date
                                                 , created_at_lte=end_date).count()
            elif start_date:
                results = Results.objects.filter(teacher=teacher, created_at__gte=start_date)
            else:
                results = Results.objects.filter(teacher=teacher).count()

            # Use FileUploadSerializer for the photo (handling None cases)
            image_data = FileUploadSerializer(teacher.photo,
                                              context={"request": request}).data if teacher.photo else None

            teacher_data.append({
                "id": teacher.id,
                "full_name": teacher.name,
                "image": image_data,
                "overall_point": teacher.overall_point,
                "subjects": subjects,
                "results": results,
            })

        return Response(teacher_data)


class GenerateExcelView(APIView):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Monitoring Table"

        # 1. Merge and style header
        ws.merge_cells('A1:A3')
        ws.merge_cells('B1:N1')
        ws['A1'] = "ASOSLAR"
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Then fill merged headings manually
        ws['B1'] = "OYLIK HISOBDA"

        # Other headers manually or in a loop
        headers = [
            "USTOZ", "MAX BALL",  # continue all the way to JAMI
            #...
        ]

        ws.append(headers)

        # 2. Set column widths
        for col in range(1, len(headers)+1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 18

        # 3. Fetch teachers and their monitoring scores
        teachers = CustomUser.objects.filter(role="TEACHER")

        for teacher in teachers:
            # calculate all points manually
            max_ball = 15  # example
            ball_plus_40 = 40  # example
            ball_plus_1000 = 1000  # etc.

            total_score = ball_plus_40 + ball_plus_1000  # etc.

            row = [
                teacher.full_name,
                max_ball,
                ball_plus_40,
                ball_plus_1000,
                # etc
                total_score
            ]
            ws.append(row)

        # 4. Set color fills (optional for columns)
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws["N"]:  # example for yellow JAMI column
            cell.fill = fill

        # 5. Return response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="monitoring_table.xlsx"'
        wb.save(response)
        return response


class DashboardWeeklyFinanceAPIView(APIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # Extract query parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        casher_id = request.query_params.get('casher')
        action_type = request.query_params.get('action')  # INCOME or EXPENSE
        kind = request.query_params.get('kind')  # Dynamically fetched kinds
        filial = request.query_params.get('filial')

        filters = {}

        # Get all dynamic kinds from DB
        existing_kinds = list(Kind.objects.values_list('name', flat=True))  # Fetch all available kinds dynamically

        if filial:
            filters["filial"] = filial
        if casher_id:
            filters['casher__id'] = casher_id

        # Ensure valid action type (INCOME or EXPENSE)
        if action_type and action_type.upper() in ["INCOME", "EXPENSE"]:
            filters['action__exact'] = action_type.upper()
        elif action_type:
            return Response({"error": "Invalid action. Use 'INCOME' or 'EXPENSE'."}, status=400)

        # Filter by start_date and end_date
        if start_date:
            try:
                filters['created_at__gte'] = datetime.strptime(start_date, '%Y-%m-%d')
            except ValueError:
                return Response({"error": "Invalid start_date format. Use YYYY-MM-DD."}, status=400)

        if end_date:
            try:
                filters['created_at__lte'] = datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return Response({"error": "Invalid end_date format. Use YYYY-MM-DD."}, status=400)

        # Ensure kind filtering is valid
        if kind:
            if kind not in existing_kinds:
                return Response({"error": f"Invalid kind. Allowed values: {', '.join(existing_kinds)}"}, status=400)
            filters['kind__name'] = kind  # Ensure kind filtering is correctly applied

        # Query and group by weekday
        queryset = (
            Finance.objects.filter(**filters)
            .annotate(weekday=ExtractWeekDay('created_at'))
            .values('weekday', 'kind__name')
            .annotate(total_amount=Sum('amount'))
        )

        # Map weekday numbers to names
        weekday_map = {
            1: 'Sunday', 2: 'Monday', 3: 'Tuesday', 4: 'Wednesday',
            5: 'Thursday', 6: 'Friday', 7: 'Saturday'
        }

        # Initialize totals for each weekday and category
        weekly_data = {day: {k: 0 for k in existing_kinds} for day in weekday_map.values()}
        total_overall = {k: 0 for k in existing_kinds}  # Overall total per category

        # Populate data
        for item in queryset:
            weekday_name = weekday_map[item['weekday']]
            category = item['kind__name']
            amount = item['total_amount']
            weekly_data[weekday_name][category] += amount
            total_overall[category] += amount  # Aggregate total for percentage calculation

        # Compute category-wise percentages
        total_sum = sum(total_overall.values())  # Total amount of all categories combined

        if total_sum > 0:
            percentages = {
                category: round((total / total_sum) * 100, 2) for category, total in total_overall.items()
            }
        else:
            percentages = {category: 0 for category in existing_kinds}

        # Convert data to response format
        result = [{"weekday": day, "totals": totals} for day, totals in weekly_data.items()]

        return Response({
            "weekly_data": result,
            "percentages": percentages,
            "available_kinds": existing_kinds  # Returning available kinds for front-end usage
        }, status=200)


class ArchivedView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        filial = request.query_params.get('filial', None)

        lid = Lid.objects.filter(is_archived=True, lid_stage_type="NEW_LID", is_student=False).count()
        order = Lid.objects.filter(is_archived=True, lid_stage_type="ORDERED_LID", is_student=False).count()
        new_student = Student.objects.filter(student_stage_type="NEW_STUDENT", is_archived=True).count()
        student = Student.objects.filter(student_stage_type="ACTIVE_STUDENT", is_archived=True).count()

        if filial:
            lid = Lid.objects.filter(is_archived=True, lid_stage_type="NEW_LID", is_student=False,filial_id=filial).count()
            order = Lid.objects.filter(is_archived=True, lid_stage_type="ORDERED_LID", is_student=False,filial_id=filial).count()
            new_student = Student.objects.filter(student_stage_type="NEW_STUDENT", is_archived=True,filial_id=filial).count()
            student = Student.objects.filter(student_stage_type="ACTIVE_STUDENT", is_archived=True,filial_id=filial).count()

        return Response({
            "lid": lid,
            "order": order,
            "new_student": new_student,
            "student": student,
        })


class SalesApiView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        creator = request.query_params.get('creator')
        sale = request.query_params.get('sale')
        student = request.query_params.get('student')
        filial = request.query_params.get('filial')

        filters = {}
        if creator:
            filters['creator_id'] = creator  # Use `_id` for ForeignKeys
        if sale:
            filters['sale_id'] = sale
        if student:
            filters['student_id'] = student
        if filial:
            filters['filial_id'] = filial

        chart_data = []

        total_students = SaleStudent.objects.filter(**filters).count()
        total_voucher_amount = \
        VoucherStudent.objects.filter(**filters).aggregate(total=Sum('voucher__amount'))['total'] or 0
        total_sale_discount = \
        SaleStudent.objects.filter(**filters).aggregate(total=Sum('sale__amount'))['total'] or 0
        total_debt = Student.objects.filter(balance__lt=0, **filters).aggregate(total=Sum('balance'))['total'] or 0

        # FIX: Removed incorrect `balance__status` lookup
        total_income = Student.objects.filter(**filters).aggregate(total=Sum('balance'))['total'] or 0

        chart_data.append({
            "total_students": total_students,
            "total_voucher_amount": total_voucher_amount,
            "total_sales_amount": total_sale_discount,
            "total_debt": total_debt,
            "active_students_balance": total_income,  # Active Students
        })

        return Response({"chart_data": chart_data})


class FinanceStatisticsApiView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        kind_id = self.request.query_params.get('kind')  # Filter by Kind
        creator = self.request.query_params.get('creator')
        student = self.request.query_params.get('student')
        stuff = self.request.query_params.get('stuff')
        casher = self.request.query_params.get('casher')
        payment_method = self.request.query_params.get('payment_method')
        filial = self.request.query_params.get('filial')

        filters = {}
        if filial:
            filters['filial'] = filial
        if kind_id:
            filters['kind__id'] = kind_id
        if creator:
            filters['creator__id'] = creator
        if student:
            filters['student__id'] = student
        if stuff:
            filters['stuff__id'] = stuff
        if casher:
            filters['casher__id'] = casher
        if payment_method:
            filters['payment_method'] = payment_method

        # ✅ Group by `kind__action` instead of `Finance.action`
        finance_stats = (
            Finance.objects.filter(**filters)
            .values("kind__id", "kind__name", "kind__action")  # Using kind__action
            .annotate(total_amount=Sum("amount"))
            .order_by("kind__name")
        )

        # ✅ Organizing INCOME and EXPENSE based on `kind__action`
        income_stats = []
        expense_stats = []
        for entry in finance_stats:
            stat_entry = {
                "kind_id": entry["kind__id"],
                "kind_name": entry["kind__name"],
                "action": entry["kind__action"],
                "total_amount": entry["total_amount"],
            }
            if entry["kind__action"] == "INCOME":  # ✅ Use `kind__action`
                income_stats.append(stat_entry)
            else:
                expense_stats.append(stat_entry)

        return Response(
            {
                "income": income_stats,
                "expenses": expense_stats,
            }
        )


class StudentLanguage(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        filial = self.request.query_params.get('filial')
        if filial:
            student_uz = Student.objects.filter(education_lang="UZB", student_stage_type = "ACTIVE_STUDENT",
                                                is_archived=False,  is_frozen=False).count()
            student_eng = Student.objects.filter(education_lang="ENG", student_stage_type = "ACTIVE_STUDENT",
                                                 is_archived=False,  is_frozen=False).count()
            student_ru = Student.objects.filter(education_lang="RU", student_stage_type = "ACTIVE_STUDENT",
                                                is_archived=False,  is_frozen=False).count()
        else:
            student_uz = Student.objects.filter(education_lang="UZB", student_stage_type = "ACTIVE_STUDENT",
                                                is_archived=False,  is_frozen=False).count()
            student_eng = Student.objects.filter(education_lang="ENG", student_stage_type = "ACTIVE_STUDENT",
                                                 is_archived=False,  is_frozen=False).count()
            student_ru = Student.objects.filter(education_lang="RU", student_stage_type = "ACTIVE_STUDENT",
                                                is_archived=False,  is_frozen=False).count()

        return Response({
            "student_uz": student_uz,
            "student_eng": student_eng,
            "student_ru": student_ru,
        })


class ExportDashboardToExcelAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # ✅ 1. Fetch Sales Data
        creator = request.query_params.get('creator')
        sale = request.query_params.get('sale')
        student = request.query_params.get('student')
        filial = request.query_params.get('filial')

        sales_filters = {}
        if creator:
            sales_filters['creator__id'] = creator
        if sale:
            sales_filters['sale__id'] = sale
        if student:
            sales_filters['student__id'] = student
        if filial:
            sales_filters['filial__id'] = filial

        student_count = SaleStudent.objects.filter(**sales_filters).values('creator').annotate(
            total_students=Count('student', distinct=True)
        )

        voucher_sales = SaleStudent.objects.filter(sale__type="VOUCHER", **sales_filters).values('creator').annotate(
            total_voucher_amount=Sum('sale__amount')
        )

        sale_data = SaleStudent.objects.filter(sale__type="SALE", **sales_filters).values('creator').annotate(
            total_sale_discount=Sum(
                Case(
                    When(
                        student__students_group__group__price_type="monthly",
                        then=F('student__students_group__group__price') * F('sale__amount') / Value(100)
                    ),
                    default=Value(0),
                    output_field=DecimalField()
                )
            )
        )

        total_sales = SaleStudent.objects.filter(**sales_filters).values('creator').annotate(
            total_sales_amount=Sum('sale__amount')
        )

        # ✅ 2. Fetch Finance Statistics Data
        finance_filters = {}
        kind_id = request.query_params.get('kind')
        action = request.query_params.get('action')
        creator = request.query_params.get('creator')
        student = request.query_params.get('student')
        stuff = request.query_params.get('stuff')
        casher = request.query_params.get('casher')
        payment_method = request.query_params.get('payment_method')

        if kind_id:
            finance_filters['kind__id'] = kind_id
        if action:
            finance_filters['action'] = action
        if creator:
            finance_filters['creator__id'] = creator
        if student:
            finance_filters['student__id'] = student
        if stuff:
            finance_filters['stuff__id'] = stuff
        if casher:
            finance_filters['casher__id'] = casher
        if payment_method:
            finance_filters['payment_method'] = payment_method

        finance_stats = Finance.objects.filter(**finance_filters).values('kind__id', 'kind__name', 'action').annotate(
            total_amount=Sum('amount')
        ).order_by('kind__name')

        # ✅ 3. Fetch Weekly Finance Data
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        casher_id = request.query_params.get('casher')
        kind = request.query_params.get('kind')

        weekly_filters = {}
        if casher_id:
            weekly_filters['casher__id'] = casher_id
        if kind:
            weekly_filters['kind__name'] = kind
        if start_date:
            weekly_filters['created_at__gte'] = start_date
        if end_date:
            weekly_filters['created_at__lte'] = end_date

        weekly_finance_data = (
            Finance.objects.filter(**weekly_filters)
            .values('kind__name')
            .annotate(total_amount=Sum('amount'))
        )

        # ✅ Create Excel Workbook
        workbook = Workbook()

        # ✅ Create Sales Sheet
        sales_sheet = workbook.active
        sales_sheet.title = "Sales Data"
        sales_headers = ["Yaratuvchi xodim", "Jami Studentlar", "Voucher Sales", "Sale Discount", "Total Sales"]
        sales_sheet.append(sales_headers)

        for entry in student_count:
            creator_id = entry['creator']
            voucher_entry = next((v for v in voucher_sales if v['creator'] == creator_id), {'total_voucher_amount': 0})
            sale_entry = next((s for s in sale_data if s['creator'] == creator_id), {'total_sale_discount': 0})
            total_sales_entry = next((t for t in total_sales if t['creator'] == creator_id), {'total_sales_amount': 0})

            sales_sheet.append([
                entry['total_students'],
                voucher_entry['total_voucher_amount'],
                sale_entry['total_sale_discount'],
                total_sales_entry['total_sales_amount'],
            ])

        # ✅ Create Finance Statistics Sheet
        finance_sheet = workbook.create_sheet(title="Finance Statistics")
        finance_headers = ["Kind Name", "Action", "Total Amount"]
        finance_sheet.append(finance_headers)

        for entry in finance_stats:
            finance_sheet.append([
                entry['kind__name'],
                entry['action'],
                entry['total_amount']
            ])

        # ✅ Create Weekly Finance Data Sheet
        weekly_sheet = workbook.create_sheet(title="Weekly Finance")
        weekly_headers = ["Kind Name", "Total Amount"]
        weekly_sheet.append(weekly_headers)

        for entry in weekly_finance_data:
            weekly_sheet.append([
                entry['kind__name'],
                entry['total_amount']
            ])

        # ✅ Style headers
        for sheet in [sales_sheet, finance_sheet, weekly_sheet]:
            for col_num, header in enumerate(sheet[1], 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                sheet.column_dimensions[cell.column_letter].width = 20

        # ✅ Create HTTP Response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="Dashboard_Data.xlsx"'
        workbook.save(response)

        return response


class AdminLineGraph(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        call_operator = request.query_params.get('call_operator')
        lid_type = request.query_params.get('type')

        # Parse date strings into actual date objects
        start_date = parse_date(start_date) if start_date else None
        end_date = parse_date(end_date) if end_date else None

        filial = request.query_params.get('filial')

        # Base filters
        lid_filter = {}
        student_filter = {}
        if filial:
            lid_filter['filial_id'] = filial
            student_filter['filial_id'] = filial

        if start_date:
            lid_filter['created_at__gte'] = start_date
            student_filter['created_at__gte'] = start_date
        if end_date:
            lid_filter['created_at__lte'] = end_date
            student_filter['created_at__lte'] = end_date
        if call_operator:
            lid_filter['call_operator'] = call_operator
        if lid_type:
            lid_filter['type'] = lid_type

        # Query for each type
        lid_new = Lid.objects.filter(lid_stage_type="NEW_LID", **lid_filter)
        lid_ordered = Lid.objects.filter(lid_stage_type="ORDERED_LID", **lid_filter)
        student_new = Student.objects.filter(student_stage_type="NEW_STUDENT", **student_filter)
        student_active = Student.objects.filter(student_stage_type="ACTIVE_STUDENT", **student_filter)

        # Function to process data and fill missing dates
        def get_counts(queryset, date_field="created_at"):
            counts = queryset.values(f"{date_field}__date").annotate(count=Count("id"))
            date_counts = defaultdict(int)

            if start_date and end_date:
                current_date = start_date
                while current_date <= end_date:
                    date_counts[current_date] = 0
                    current_date += timedelta(days=1)

            for entry in counts:
                date_counts[entry[f"{date_field}__date"]] = entry["count"]

            return [{"date": date.strftime("%Y-%m-%d"), "count": count} for date, count in sorted(date_counts.items())]

        # Construct the response
        response_data = {
            "new_lid": (get_counts(lid_new)),
            "ordered_lid": get_counts(lid_ordered),
            "new_student": get_counts(student_new),
            "active_student": get_counts(student_active),
        }

        return Response(response_data)