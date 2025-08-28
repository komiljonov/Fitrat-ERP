import logging
import os
import pandas as pd
from celery import shared_task
from datetime import datetime, timedelta
from django.conf import settings
from aiogram import Bot
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from requests import post

from data.account.models import CustomUser
from data.department.filial.models import Filial
from data.finances.finance.models import Finance
from data.lid.new_lid.models import Lid
from data.student.groups.models import Group
from data.student.lesson.models import FirstLLesson
from data.student.student.models import Student
from data.student.studentgroup.models import StudentGroup

logging.basicConfig(level=logging.INFO)

class TelegramBot:
    HOST = "https://api.telegram.org/bot"

    def __init__(self):
        # Get token from environment variables


        token = os.getenv("BOT_TOKEN")
        if not token:
            raise ValueError("Telegram bot TOKEN is missing! Set the TOKEN environment variable.")
        self.base_url = self.HOST + token

    def send_message(
        self,
        chat_id,
        text,
        reply_markup=None,
        parse_mode="HTML",
    ):
        """
        Sends a message via Telegram Bot API.
        """
        url = self.base_url + "/sendMessage"
        data = {
            "chat_id": chat_id,
            "text": text,
            "parse_mode": parse_mode,
        }

        # Add reply markup if provided
        if reply_markup:
            data["reply_markup"] = reply_markup.to_json()

        # Send the request
        try:
            res = post(url, json=data)
            res.raise_for_status()  # Raise HTTP errors if they occur
            logging.info(f"Message sent to chat_id {chat_id}: {text}")
            return res.json()
        except Exception as e:
            logging.error(f"Failed to send message to chat_id {chat_id}: {e}")
            return None

    def send_document(self, chat_id, file_path, caption="📄 Here is your document"):
        """
        Sends a document to a Telegram user.
        :param chat_id: Telegram chat ID of the recipient
        :param file_path: Path to the document file
        :param caption: Optional caption for the document
        """
        url = self.base_url + "/sendDocument"
        try:
            with open(file_path, "rb") as document:
                files = {"document": document}
                data = {"chat_id": chat_id, "caption": caption}
                res = post(url, data=data, files=files)
                res.raise_for_status()
                logging.info(f"Document sent to chat_id {chat_id}: {file_path}")
                return res.json()
        except Exception as e:
            logging.error(f"Failed to send document to chat_id {chat_id}: {e}")
            return None

bot = TelegramBot()

def generate_excel_report():
    """
    Generates an Excel report with student data on the first sheet
    and financial transactions on the second sheet.
    """
    # Generate filename
    file_name = f"report_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)

    # Create a new workbook and sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Kunlik reportlar"

    # 🔹 Header Row for Student Data
    headers = ["Jarayonlar"]
    filials = [filial.name for filial in Filial.objects.all()]
    headers += filials + ["Jami"]  # Add "Total" column

    # 🔹 Initialize Data Collection
    data_dict = {
        "Buyurtma": [],
        "Birinchi darsga keladiganlar": [],
        "Yangi o‘quvchi": [],
        "Aktiv": [],
        "Qarzdorlar": [],
        "Guruhlar": [],
        "Shu oyda to‘lov qilganlar": [],
        "Arxiv": [],
        "JAMI REAL BOR": [],
        "Jami o‘quvchi": [],
        "Jami aktiv": [],
        "Guruhdagi jami aktiv o‘quvchilar": [],
    }

    for filial in Filial.objects.all():
        # 🔹 Fetch student data
        orders = Lid.objects.filter(filial=filial, ordered_date__gte=datetime.now() - timedelta(days=1), is_archived=False).count()
        first_lesson = FirstLLesson.objects.filter(filial=filial, created_at__gte=datetime.now() - timedelta(days=1), lid__is_archived=False).count()
        new_student = Student.objects.filter(filial=filial, student_stage_type="NEW_STUDENT",
                                             new_student_date__gte=datetime.now() - timedelta(days=1), is_archived=False).count()
        student = Student.objects.filter(filial=filial, student_stage_type="ACTIVE_STUDENT",
                                         active_date__gte=datetime.now() - timedelta(days=1), is_archived=False).count()
        debt = Student.objects.filter(filial=filial, balance_status="INACTIVE", is_archived=False).count()
        groups = Group.objects.filter(filial=filial, status="ACTIVE").count()
        finance_student = Finance.objects.filter(student__isnull=False, filial=filial, kind__name="Course payment",
                                                 created_at__lte=datetime.today().replace(day=1), created_at__gte=datetime.today()).count()
        archived_lid = Lid.objects.filter(filial=filial, is_student=False, is_archived=True).count()
        archived_student = Student.objects.filter(filial=filial, is_archived=True).count()
        all_archived = archived_lid + archived_student
        all_real_students = Student.objects.filter(filial=filial, is_archived=False, is_frozen=False).count()
        all_students = StudentGroup.objects.filter(filial=filial, group__status="ACTIVE", student__is_archived=False).count()
        all_active = Student.objects.filter(filial=filial, balance_status="ACTIVE").count()
        groups_active_students = StudentGroup.objects.filter(filial=filial, group__status="ACTIVE", student__is_archived=False).count()

        # 🔹 Store in Data Dictionary
        data_dict["Buyurtma"].append(orders)
        data_dict["Birinchi darsga keladiganlar"].append(first_lesson)
        data_dict["Yangi o‘quvchi"].append(new_student)
        data_dict["Aktiv"].append(student)
        data_dict["Qarzdorlar"].append(debt)
        data_dict["Guruhlar"].append(groups)
        data_dict["Shu oyda to‘lov qilganlar"].append(finance_student)
        data_dict["Arxiv"].append(all_archived)
        data_dict["JAMI REAL BOR"].append(all_real_students)
        data_dict["Jami o‘quvchi"].append(all_students)
        data_dict["Jami aktiv"].append(all_active)
        data_dict["Guruhdagi jami aktiv o‘quvchilar"].append(groups_active_students)

    # 🔹 Calculate Totals for each row
    for key in data_dict.keys():
        data_dict[key].append(sum(data_dict[key]))

    # 🔹 Convert data dictionary to a list of rows
    data = [[key] + values for key, values in data_dict.items()]

    # 🔹 Styling
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, size=12)
    title_fill = PatternFill("solid", fgColor="4CAF50")
    header_fill = PatternFill("solid", fgColor="FFFF00")
    red_fill = PatternFill("solid", fgColor="FF0000")
    green_fill = PatternFill("solid", fgColor="00FF00")
    cyan_fill = PatternFill("solid", fgColor="00FFFF")

    # 🔹 Merge and style title row
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws1["A1"] = datetime.now().strftime("%d/%m/%Y")
    ws1["A1"].font = title_font
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = Alignment(horizontal="center")

    # 🔹 Insert Headers
    ws1.append(headers)
    for col_idx in range(2, len(headers) + 1):
        ws1[f"{get_column_letter(col_idx)}2"].font = header_font
        ws1[f"{get_column_letter(col_idx)}2"].fill = header_fill
        ws1[f"{get_column_letter(col_idx)}2"].alignment = Alignment(horizontal="center")

    # 🔹 Insert Student Data
    for row in data:
        ws1.append(row)
        for col_idx, value in enumerate(row[1:], start=2):
            cell = ws1.cell(row=ws1.max_row, column=col_idx)
            if "Qarzdorlar" in row[0]:  # Red fill for debts
                cell.fill = red_fill
            elif "Aktiv" in row[0]:  # Green fill for active students
                cell.fill = green_fill
            elif "JAMI REAL BOR" in row[0]:  # Yellow fill for summary
                cell.fill = header_fill
            elif "Jami o‘quvchi" in row[0]:  # Cyan fill for total students
                cell.fill = cyan_fill

    # 🔹 Create Second Sheet for Financial Transactions
    ws2 = wb.create_sheet(title="Moliya Hisobot")

    finance_headers = ["Turi", "Kim", "Comment", "Qiy mati", "To'lov turi", "Kassa"]
    ws2.append(finance_headers)

    finance_entries = Finance.objects.all()

    for entry in finance_entries:
        ws2.append([
            "Student to'ladi" if entry.student else "Hodim avans" if entry.stuff else "Boshqa",
            entry.student.first_name + " " + entry.student.last_name if entry.student else entry.stuff.full_name if entry.stuff else "Unknown",
            entry.comment or "-",
            entry.amount,
            entry.payment_method,
            entry.casher.user.full_name if entry.casher else "-"
        ])

    # 🔹 Auto adjust column width for both sheets
    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 3

    # 🔹 Save the Excel file
    wb.save(file_path)
    return file_path

@shared_task
def send_daily_excel_report():
    """
    Celery task to generate and send the Excel report via Telegram.
    """
    logging.info("Generating Excel report...")

    # Generate the Excel report
    file_path = generate_excel_report()

    if not os.path.exists(file_path):
        logging.error("Excel report was not generated correctly.")
        return "Failed to generate the report."

    users = CustomUser.objects.filter(chat_id__isnull=False, role__in=["DIRECTOR", "MULTIPLE_FILIAL_MANAGER"])

    success_count = 0

    for user in users:
        try:
            bot.send_document(chat_id=user.chat_id, file_path=file_path, caption="📊 Kunlik hisobot")
            logging.info(f"✅ Report sent to {user.first_name} ({user.chat_id})")
            success_count += 1
        except Exception as e:
            logging.error(f"❌ Error sending report to {user.chat_id}: {str(e)}")

    # Cleanup: Remove the temporary file
    try:
        os.remove(file_path)
        logging.info("Temporary Excel report deleted.")
    except Exception as e:
        logging.warning(f"Failed to delete report file: {str(e)}")

    logging.info(f"✅ Celery task completed: Sent Excel report to {success_count} users.")
    return f"Excel report sent to {success_count} users."